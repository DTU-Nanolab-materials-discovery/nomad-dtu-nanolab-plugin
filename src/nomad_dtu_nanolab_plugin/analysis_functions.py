# %%
import os
import pickle
import re

import matplotlib.colors as mcolors
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go

#from lmfit import Parameters
from lmfit.models import LinearModel, PseudoVoigtModel  #, SplineModel, GaussianModel,
from openpyxl import load_workbook
from plotly.subplots import make_subplots
from scipy.interpolate import griddata
from scipy.signal import find_peaks, savgol_filter

##########################
# Functions related to loading data
##########################

def convert_to_eV(data):
    """Convert ellipsometry data in wavelength to eV."""
    # Planck's constant (eV/Hz)
    h = 4.135 * 10**-15
    # Speed of light (m/s)
    c = 3 * 10**8
    data = data.copy()
    data.iloc[:, data.columns.get_level_values(1) == 'Wavelength (nm)'] = (h * c) / (
        data.iloc[:, data.columns.get_level_values(1) == 'Wavelength (nm)'] * 10**-9
    )
    data.columns = data.columns.set_levels(['Energy (eV)', 'k', 'n'], level=1)
    data = data.round(3)
    return data

##########################
# Functions related to constructing a grid and inserting data into it
##########################


def measurement_grid(
    ncolumns, nrows, gridlength, gridheight, startposition= [0,0]
):
    """Define a grid based on number of columns and rows, length and height of grid in
    mm, and the first coordinate (lower left corner) in the column and row."""
    startcolumn = startposition[0]
    startrow = startposition[1]

    xgrid = np.round(np.linspace(startcolumn, gridlength + startcolumn, ncolumns), 3)
    ygrid = np.round(np.linspace(startrow, gridheight + startrow, nrows), 3)
    grid = np.array([xgrid[0], ygrid[0]])
    for i in range(len(xgrid)):
        for j in range(len(ygrid)):
            grid = np.vstack((grid, np.array([xgrid[i], ygrid[j]])))
    grid = grid[1:]
    grid = pd.DataFrame(grid, columns=['x', 'y'])
    return grid


def coords_to_grid(coords, grid):
    """Constrain a set of measured datapoints to a custom defined grid made with the
    "measurement_grid" function."""
    griddata = coords.copy()
    for i in range(len(coords)):
        # find closest x and y coordinate
        xminindex = np.abs(grid - coords.iloc[i, :]).idxmin().iloc[0]
        yminindex = np.abs(grid - coords.iloc[i, :]).idxmin().iloc[1]
        # assign new coordinates
        griddata.iloc[i, 0] = np.round(grid.iloc[xminindex, 0], 2)
        griddata.iloc[i, 1] = np.round(grid.iloc[yminindex, 1], 2)
    return griddata


def grid_to_MIheader(grid):
    """Convert a grid (array of x,y) into a multi index header"""
    MIgrid = []
    for i in range(len(grid)):
        MIgrid = np.append(MIgrid, (f'{grid.iloc[i,0]},{grid.iloc[i,1]}'))
    return MIgrid


def MI_to_grid(MIgrid):
    """Convert multi index into a grid (array of x,y)"""
    MIgrid = MIgrid.columns.get_level_values(0)
    splitvals = re.split(',', MIgrid[0])
    grid = np.array([splitvals[0], splitvals[1]])
    for i in range(1, len(MIgrid)):
        splitvals = re.split(',', MIgrid[i])
        grid = np.vstack((grid, np.array([splitvals[0], splitvals[1]])))
    grid = pd.DataFrame(grid, columns=['x', 'y'])
    grid = grid.astype(float)
    return grid


def interpolate_grid(data, grid):
    """Interpolate data over a custom grid made with the "measurement_grid"
    function."""
    # !!!
    # specifically remove the "Peak" column, which will be present if loaded data is XPS
    # !!!
    data = data.drop(columns='Peak', level=1, errors='ignore')

    # get grid-aligned coordinates for datapoints for interpolation
    coords = MI_to_grid(data).drop_duplicates(ignore_index=True)

    # interpolation
    # we have to account for multiple variables for every coordinate
    # this creates a list of the data
    dataT = data.transpose()
    dataN = int(len(dataT) / len(coords))
    interpolated_data = []
    for i in range(dataN):
        interpolated_data.append(
            griddata(coords, dataT.iloc[i::dataN], grid, method='cubic')
        )

    # convert the list of data to an array with columns alternating between data types
    if dataN == 0:
        interpolated_array = np.transpose(interpolated_data)
    else:
        interpolated_array = [None] * len(interpolated_data[0]) * dataN
        for i in range(dataN):
            interpolated_array[i::dataN] = interpolated_data[i]
        interpolated_array = np.transpose(interpolated_array)

    # list of column names
    columnlist = []
    for i in range(interpolated_array.shape[1]):
        columnlist.append(data.columns[i % dataN][1])

    # dataframe
    interpolated_frame = pd.DataFrame(interpolated_array, columns=columnlist)

    # convert grid to multiindex header
    coord_header = grid_to_MIheader(grid)

    # construct dataframe with multiindexing for coordinates
    header = pd.MultiIndex.from_product(
        [coord_header, list(interpolated_frame.columns.unique())],
        names=['Coordinate', 'Data type'],
    )
    interpolated_frame = pd.DataFrame(interpolated_frame.values, columns=header)
    return interpolated_frame

def extract_coordinates(data):
    coords = data.columns.get_level_values(0).unique().values
    x_values = []
    y_values = []

    for item in coords:
        x, y = item.split(',')
        x_values.append(float(x))
        y_values.append(float(y))

    return x_values, y_values

def snake_grid(x, y):  # x and y are lists of coordinates you should take note of
    X_snake = []
    Y_snake = []

    # Loop through each y-coordinate from bottom to top
    for i, y_val in enumerate(y):
        if i % 2 == 0:
            X_snake.extend(x)  # Even row: left to right ( add x normally)
        else:
            X_snake.extend(x[::-1])  # Odd row: right to left (add x in reverse)
        Y_snake.extend([y_val] * len(x))  # add as many y values as x values

    grid_snake = pd.DataFrame({'x': X_snake, 'y': Y_snake})
    return grid_snake

def select_points(data, x_min=-40, x_max=40, y_min=-40, y_max=40):
    """get coordinates of the points within the defined range, you can call them with
    get_data, or plot_data, or interactive_XRD_shift"""
    grid0 = MI_to_grid(data)
    grid = grid0.drop_duplicates().reset_index(drop=True)

    grid1 = grid[grid['x'] > x_min]
    grid2 = grid1[grid1['x'] < x_max]
    grid3 = grid2[grid2['y'] > y_min]
    grid4 = grid3[grid3['y'] < y_max]
    new_x = grid4['x'].values
    new_y = grid4['y'].values
    return new_x, new_y

def rotate_coordinates(data_df, how='clockwise'):
    """Rotate the coordinates of the data by 90 degrees clockwise, counterclockwise or
    180 degrees"""
    MI_rotated = []
    initial_coordinates = MI_to_grid(data_df)

    if how == 'clockwise':
        xx = initial_coordinates['y']
        yy = -initial_coordinates['x']

    if how == 'counterclockwise':
        xx = -initial_coordinates['y']
        yy = initial_coordinates['x']

    if how == '180':
        xx = -initial_coordinates['x']
        yy = -initial_coordinates['y']

    for i in range(len(xx)):
        MI_rotated = np.append(MI_rotated, (f'{xx[i]},{yy[i]}'))
    rotated_columns = pd.MultiIndex.from_tuples(
        [
            (str(coord), col)
            for coord, col in zip(MI_rotated, data_df.columns.get_level_values(1))
        ]
    )
    data_rotated = data_df.copy()
    data_rotated.columns = rotated_columns
    return data_rotated


##########################
# Functions related to combining, saving, and loading data and helper functions
##########################

def add_info(data, info_dict):
    """Function to add information to a dataset for each point."""
    info_type = list(info_dict.keys())[0]
    headerlength = len(data.columns.get_level_values(1).unique())
    coords = data.columns.get_level_values(0).unique()
    k = 0
    new_data = data.copy()
    for i in range(0, len(coords)):
        # print(coords[i])
        new_df = pd.DataFrame(
            [info_dict[info_type][i]], columns=[(coords[i], info_type)]
        )

        new_data.insert(
            headerlength * (i + 1) + k,
            f'{data.columns.get_level_values(0).unique()[i]}',
            new_df,
            allow_duplicates=True,
        )
        new_data.rename(columns={'': f'{info_type}'}, inplace=True)

        k = k + len(new_df.columns)

    new_frame = new_data.copy()

    return new_frame

def closest_coord(grid, x, y):
    """Find closest x and y coordinate for a grid."""
    xminindex = np.abs(grid - x).idxmin().iloc[0]
    xcoord = grid.iloc[xminindex, 0]
    yminindex = np.abs(grid[grid['x'] == xcoord] - y).idxmin().iloc[1]
    ycoord = grid.iloc[yminindex, 1]
    return xcoord, ycoord


def combine_data(datalist):
    """Combine multiple measurements into a single dataframe."""
    dataframe = pd.concat(datalist, axis=1)
    return dataframe

def math_on_columns(data, type1, type2, operation='/'):
    """Perform an operation on two columns in a provided dataframe. Usage:
    math_on_columns(data, datatype1, datatype2, operation), where "operation" can be
    +, -, *, or /."""
    coordinatelength = len(data.columns.get_level_values(0).unique())
    headerlength = len(data.columns.get_level_values(1).unique())
    k = 0
    # do math on values
    data = data.copy()
    for i in range(coordinatelength):
        val1 = data.iloc[:, data.columns.get_level_values(1) == type1].iloc[:, i]
        if isinstance(type2, str):
            val2 = data.iloc[:, data.columns.get_level_values(1) == type2].iloc[:, i]
        if isinstance(type2, (int, float)):
            val2 = type2
        if operation == '+':
            resultval = val1 + val2
        elif operation == '-':
            resultval = val1 - val2
        elif operation == '*':
            resultval = val1 * val2
        elif operation == '/':
            try:
                resultval = val1 / val2
            except ZeroDivisionError:
                resultval = float('NaN')
        # insert result
        data.insert(
            headerlength * (i + 1) + k,
            f'{data.columns.get_level_values(0).unique()[i]}',
            resultval,
            allow_duplicates=True,
        )
        k += 1

    # rename added columns
    if operation == '+':
        rowname = f'{type1} + {type2}'
    elif operation == '-':
        rowname = f'{type1} - {type2}'
    elif operation == '*':
        rowname = f'{type1} * {type2}'
    elif operation == '/':
        rowname = f'{type1} / {type2}'
    data.rename(columns={'': rowname}, inplace=True)
    return data


def get_data(data, type='all', x='all', y='all', drop_nan=True):
    """Print a data type from a multi index dataframe at a specific coordinate. The
    coordinate does not have to be exact. Leave type as blank or 'all' to select all
    types. Leave coordinates blank or 'all' to select all coordinates."""
    if x == 'all' and y == 'all':
        if type == 'all':
            print('All data at all coordinates.')
            if drop_nan is True:
                data = data.dropna(axis=0, how='all').fillna('-')
            return data
        else:
            print(f'{type} data at all coordinates.')
            if drop_nan is True:
                data = data.dropna(axis=0, how='all').fillna('-')
            return data.iloc[:, data.columns.get_level_values(1) == type]
    else:
        datagrid = MI_to_grid(data)
        # find closest x and y coordinate
        xcoord, ycoord = closest_coord(datagrid, x, y)
        coords = f'{xcoord},{ycoord}'
        if type == 'all':
            print(f'All data at {x},{y}.')
            if drop_nan is True:
                data = data.dropna(axis=0, how='all').fillna('-')
            return data.xs(coords, axis=1)
        else:
            print(f'{type} data at {x},{y}.')
            if drop_nan is True:
                data = data.dropna(axis=0, how='all').fillna('-')
            return data.xs(coords, axis=1)[type]


def translate_data(data, x, y):
    """Move a set of datapoints by a given x and y offset. Useful when combining
    multiple samples into one dataframe."""
    coords = MI_to_grid(data)
    coords['x'] = coords['x'] + x
    coords['y'] = coords['y'] + y
    coord_header = grid_to_MIheader(coords)
    header = pd.MultiIndex.from_arrays(
        [coord_header, data.columns.get_level_values(1)],
        names=['Coordinate', 'Data type'],
    )
    data = pd.DataFrame(data.values, columns=header)
    return data, coords

def save_data(dataframe, filename, separator='\t'):
    """Save dataframe to tab seperated txt file."""
    dataframe.to_csv(filename, separator, index=False, encoding='utf-8')


def load_data(filepath, separator='\t'):
    """Load txt to dataframe."""
    dataframe = pd.read_csv(filepath, sep=separator, header=[0, 1])
    dataframe.columns.rename(['Coordinate', 'Data type'], level=[0, 1], inplace=True)
    return dataframe

def export_specific(data, type, x, y, path):
    'export a specific point in XY format in a .txt file'
    data_exp = get_data(data, type=type, x=x, y=y)
    data_exp.to_csv(path, sep='\t', index=False, header=False)
    print(data_exp)

#########################
# Functions related to EDX analysis, layerprobe
#########################

def EDX_stage_coords(folder, filename):
    """Calculate EDX coordinates for a given file. Requires .xlsx file with columns as
    in template."""

    # Define file paths
    filepath = os.path.join(folder, filename + '.xlsx')
    newpath = os.path.join(folder, filename + '_stage_coords.xlsx')

    # Read specific columns from the Excel file, and extract values
    file = pd.read_excel(filepath, sheet_name='Sheet2', usecols='H:P')
    file = file.drop(
        file.index[3:]
    )  # Drop the 4th row (assumes you have info for 3 points)

    nrows = file['nrows'].values[0].astype(int)
    ncolumns = file['ncolumns'].values[0].astype(int)
    points_x = file['points x'].values[0:3]
    points_y = file['points y'].values[0:3]

    # Calculate spacing between points 1, 3, (nrows-1)
    space_x = points_x[2] - points_x[0]
    space_y = points_y[1] - points_y[0]

    # Generate coordinates and order them as layerprobe does
    coord_x = np.round(
        np.linspace(points_x[0], points_x[0] + space_x * (ncolumns - 1), ncolumns), 2
    )
    coord_y = np.round(
        np.linspace(points_y[0], points_y[0] + space_y * (nrows - 1), nrows), 2
    )

    X, Y = [], []
    for j in range(ncolumns):
        for i in range(nrows):
            Y.append(coord_y[i])
            X.append(coord_x[j])

    # Load the workbook and insert the calculated coordinates in the first sheet
    workbook = load_workbook(filepath)

    for i, value in enumerate(X, start=2):
        workbook['Sheet1'][f'B{i}'] = value
    for i, value in enumerate(Y, start=2):
        workbook['Sheet1'][f'C{i}'] = value

    workbook.save(newpath)
    workbook.close()

    print(filename, ' - coordinates calculated and saved')


def EDX_sample_coords(folder, filename):
    """Calculate and translate EDX coordinates for a given file. Requires .xlsx file
    with columns as in template."""

    # Define file paths
    filepath = os.path.join(folder, filename + '.xlsx')
    newpath = os.path.join(folder, filename + '_sample_coords.xlsx')

    # Read specific columns from the Excel file, and extract values
    file = pd.read_excel(filepath, sheet_name='Sheet2', usecols='H:P')
    file = file.drop(
        file.index[3:]
    )  # Drop the 4th row (assumes you have info for 3 points)

    nrows = file['nrows'].values[0].astype(int)
    ncolumns = file['ncolumns'].values[0].astype(int)
    corners_x = file['corner x'].values[0:2]
    corners_y = file['corner y'].values[0:2]
    points_x = file['points x'].values[0:3]
    points_y = file['points y'].values[0:3]

    # Calculate spacing between points 1, 3, (nrows-1)
    space_x = points_x[2] - points_x[0]
    space_y = points_y[1] - points_y[0]

    # Calculate shift from corners and correct for this translation
    shift_x = (corners_x[1] + corners_x[0]) / 2
    shift_y = (corners_y[0] + corners_y[1]) / 2
    start_x = points_x[0] - shift_x
    start_y = points_y[0] - shift_y

    # Generate coordinates and order them as layerprobe does
    coord_x = np.round(
        np.linspace(start_x, start_x + space_x * (ncolumns - 1), ncolumns), 2
    )
    coord_y = np.round(np.linspace(start_y, start_y + space_y * (nrows - 1), nrows), 2)
    X, Y = [], []
    for j in range(0, ncolumns):
        for i in range(0, nrows):
            Y.append(coord_y[i])
            X.append(coord_x[j])

    # Load the workbook and insert the calculated coordinates in the first sheet
    workbook = load_workbook(filepath)
    sheet1 = workbook['Sheet1']

    for i, value in enumerate(X, start=2):
        sheet1[f'B{i}'] = value
    for i, value in enumerate(Y, start=2):
        sheet1[f'C{i}'] = value

    workbook.save(newpath)
    workbook.close()

    # check for correct translation (allow 0.3 mm misalignment from sample rotation)
    allowed_dev = 0.3
    if np.abs(X[-1] + X[0]) > allowed_dev or np.abs(Y[-1] + Y[0]) > allowed_dev:
        print(filename, ' - coordinates calculated and saved, but not symmetric')
        print('X shift: ', X[-1] + X[0])
        print('Y shift: ', Y[-1] + Y[0])
    else:
        print(filename, ' - coordinates calculated, translated and saved')


def EDX_coordinates(folder, filename, edge=3, rotate=False, spacing='auto'):
    """Calculate and translate EDX coordinates for a given file. Requires .xlsx file
    with columns as in template."""

    # Define file paths
    filepath = os.path.join(folder, filename + '.xlsx')
    newpath = os.path.join(folder, filename + '_new_coords.xlsx')

    # Read specific columns from the Excel file, and extract values
    file = pd.read_excel(filepath, sheet_name='Sheet2', usecols='H:P')
    file = file.drop(
        file.index[3:]
    )  # Drop the 4th row (assumes you have info for 3 points)

    nrows = file['nrows'].values[0].astype(int)
    ncolumns = file['ncolumns'].values[0].astype(int)
    corners_x = file['corner x'].values[0:2]
    corners_y = file['corner y'].values[0:2]
    mag = file['magnification'].values[0]
    if spacing == 'auto':
        spacing = file['spacing'].values[0]

    if rotate == '90':
        ncolumns, nrows = nrows, ncolumns
        areax = 2.8 * 100 / mag
        areay = 4.1 * 100 / mag

    if rotate is False:
        # Calculate the effective area size in the x-direction, considering
        # magnification,
        # assuming the x/y ratio is constant 4.1 : 2.8
        areax = 4.1 * 100 / mag
        areay = 2.8 * 100 / mag

    # Calculate the spacing , gridlength and starting x-coordinate for the grid in
    # x-direction (assuming the grid is centered)
    space_x = areax * spacing / 100
    gridlength = (ncolumns - 1) * (space_x + areax)  # + areax
    startx = -gridlength / 2  # + (areax / 2)

    # do the same for the y-direction

    space_y = areay * spacing / 100
    gridheight = (nrows - 1) * (space_y + areay)  # + areay/2
    starty = -gridheight / 2  # + (areay/2)

    samplesize = [corners_x[1] - corners_x[0], corners_y[0] - corners_y[1]]
    print('Sample size is', samplesize)

    # Check if the grid dimensions exceed the maximum allowed size (31x31 mm)
    # if so, reduce the spacing by 10% and try again
    if gridlength >= samplesize[0] - edge or gridheight >= samplesize[1] - edge:
        print('Spacing is too large for the map')

        new_spacing = np.round(spacing - spacing * 0.05, 0)
        print('New spacing is', new_spacing)

        return EDX_coordinates(folder, filename, spacing=new_spacing)

    # Generate coordinates for each column
    coord_x = np.round(np.linspace(startx, -startx, ncolumns), 2)
    coord_y = np.round(np.linspace(starty, -starty, nrows), 2)
    X = []
    Y = []
    for j in range(0, ncolumns):
        for i in range(0, nrows):
            Y.append(coord_y[i])
            X.append(coord_x[j])

    # Load the workbook and insert the calculated coordinates in the first sheet
    workbook = load_workbook(filepath)
    sheet1 = workbook['Sheet1']

    for i, value in enumerate(X, start=2):
        sheet1[f'B{i}'] = value
    for i, value in enumerate(Y, start=2):
        sheet1[f'C{i}'] = value

    workbook.save(newpath)
    workbook.close()


def lp_translate_excel(folder, filename):
    """Creates a new excel file with translated coordinates, given the coordinates
    of the corners in Sheet2, assuming they are stored rightafter the statistics"""
    filepath = os.path.join(folder, filename + '.xlsx')
    newpath = os.path.join(folder, filename + '_translated.xlsx')

    first_data = pd.read_excel(filepath, sheet_name='Sheet1')

    first_x = first_data['X (mm)']
    first_y = first_data['Y (mm)']

    corners = pd.read_excel(filepath, sheet_name='Sheet2', usecols='K:L')

    trans_x = corners.iloc[[0, 1], 0].mean()
    trans_y = corners.iloc[[0, 1], 1].mean()

    new_x = first_x - trans_x
    new_y = first_y - trans_y

    new_data = first_data.copy()
    new_data['X (mm)'] = new_x
    new_data['Y (mm)'] = new_y

    # new_data.to_excel(new_path, index = False)

    workbook = load_workbook(filepath)
    sheet1 = workbook['Sheet1']

    for i, value in enumerate(new_x, start=2):
        sheet1[f'B{i}'] = value
    for i, value in enumerate(new_y, start=2):
        sheet1[f'C{i}'] = value

    workbook.save(newpath)
    workbook.close()

def find_composition(
    data,
    element_list, #shape element_list = ['Si', 'Ge', 'O']
    range_array, #shape range_array = [[0, 3], [4, 7], [8, 11]]
    stoichiometry=None,
    sample='sample',
):
    """find te points in the sample where the composition is in a certain range, given
    in % ranges or in stoichiometry and tolerance"""
    el1 = element_list[0]
    el2 = element_list[1]
    el3 = element_list[2]
    range1 = [range_array[0][0], range_array[0][1]]
    range2 = [range_array[1][0], range_array[1][1]]
    range3 = [range_array[2][0], range_array[2][1]]
    tolerance=3


    if stoichiometry:
        tot = sum(stoichiometry)
        range1 = [
            (stoichiometry[0] * 100 / tot) - tolerance,
            (stoichiometry[0] * 100 / tot) + tolerance,
        ]
        range2 = [
            (stoichiometry[1] * 100 / tot) - tolerance,
            (stoichiometry[1] * 100 / tot) + tolerance,
        ]
        range3 = [
            (stoichiometry[2] * 100 / tot) - tolerance,
            (stoichiometry[2] * 100 / tot) + tolerance,
        ]

    ranges = [range1, range2, range3]
    elements = [el1, el2, el3]
    indices = [None, None, None]

    for i in range(0, len(elements)):
        idx_min = np.where(
            get_data(data, type=f'Layer 1 {elements[i]} Atomic %').values[0]
            > ranges[i][0]
        )[0]
        idx_max = np.where(
            get_data(data, type=f'Layer 1 {elements[i]} Atomic %').values[0]
            < ranges[i][1]
        )[0]
        idx = np.intersect1d(idx_max, idx_min)
        if 0 <= i < len(indices):
            indices[i] = idx
        else:
            print('Error')
    idx1 = indices[0]
    idx2 = indices[1]
    idx3 = indices[2]
    idx = np.intersect1d(idx1, idx2)
    idx = np.intersect1d(idx, idx3)
    x, y = extract_coordinates(data)
    good_comp = {'X': [], 'Y': []}
    for i in range(0, len(idx)):
        good_comp['X'].append(x[idx[i]])
        good_comp['Y'].append(y[idx[i]])

    good_comp = pd.DataFrame(good_comp)
    # display(good_comp)
    plt.scatter(good_comp['X'], good_comp['Y'], c='r', s=80)
    plt.scatter(x, y, c='b', s=10)
    plt.xlabel('x position (mm)')
    plt.ylabel('y position (mm)')
    if stoichiometry:
        plt.title(
            f'{sample} - Positions where composition is {el1}{stoichiometry[0]}, '
            f'{el2}{stoichiometry[1]}, {el3}{stoichiometry[2]} +-{tolerance}%'
        )
    else:
        plt.title(
            f'{sample} - Positions where {el1}: {range1[0]:.1f}-{range1[1]:.1f}%, '
            f'{el2}: {range2[0]:.1f}-{range2[1]:.1f}%, {el3}: {range3[0]:.1f}-'
            f'{range3[1]:.1f}%'
        )

    for i in range(0, len(good_comp)):
        print(get_data(data, x=good_comp['X'][i], y=good_comp['Y'][i]))

    return good_comp


def calculate_ratio(df, el1, el2, rename=None):
    df = math_on_columns(df, f'Layer 1 {el1} Atomic %', f'Layer 1 {el2} Atomic %', '/')
    if rename:
        df.rename(
            columns={f'Layer 1 {el1} Atomic % / Layer 1 {el2} Atomic %': rename},
            inplace=True,
        )
    else:
        df.rename(
            columns={
                f'Layer 1 {el1} Atomic % / Layer 1 {el2} Atomic %': f'{el1}/{el2}'
            },
            inplace=True,
        )
    return df


def calculate_el_thickness(df, el):
    df = math_on_columns(df, f'Layer 1 {el} Atomic %', 'Layer 1 Thickness (nm)', '*')
    df = math_on_columns(
        df, f'Layer 1 {el} Atomic % * Layer 1 Thickness (nm)', 100, '/'
    )
    df.rename(
        columns={f'Layer 1 {el} Atomic % * Layer 1 Thickness (nm) / 100': f'{el} [nm]'},
        inplace=True,
    )
    df.drop(
        columns=f'Layer 1 {el} Atomic % * Layer 1 Thickness (nm)', level=1, inplace=True
    )
    return df

def stats(data_all, type):
    data = get_data(data_all, type = type)
    data = data.sort_values(by = 0, axis=1 )
    min_= data.iloc[0,0]
    max_ = data.iloc[0,-1]
    mean_ = data.mean(axis=1)[0]

    data = pd.DataFrame([min_, max_, mean_], index = ["min","max", "mean"])
    return data


##########################
# Functions related to XRD analysis
##########################


def initial_peaks(
    data,
    dataRange,
    filterstrength,
    peakprominence,
    peakwidth,

):
    """finds peaks using scipy find_peaks on filtered data to construct a model for
    fitting, filter strength is based on filtervalue and
    peak find sensitivity based on peakprominence, withplots and plotscale allows
    toggling plots on/off and picking scale.
    Output: dataframe with peak locations and intensity, to be used for raman_fit or
    xrd_fit, and data limited by the dataRangemin/max, in index"""
    plotscale='log'
    dataRangeMin = dataRange[0]
    dataRangeMax = dataRange[1]
    peakheight=0 # not sure if good here butgood enough for now!!!
    # setup data
    column_headers = data.columns.values
    col_theta = column_headers[::2]
    col_counts = column_headers[1::2]
    data = data.iloc[dataRangeMin:dataRangeMax]
    data.reset_index(drop=True, inplace=True)

    # create list for intital peaks
    thePeakss = []
    dataCorrect1 = []

    # finding the peaks
    for i in range(0, len(col_theta)):
        # select data
        dataSelect = data[col_counts[i]].copy()
        x = data[col_theta[i]]

        # Filter to avoid fake peaks
        if filterstrength > 0:
            # dataSelect = lfilter(b, a, dataSelect)
            dataSelect = savgol_filter(dataSelect, filterstrength, 1)

        # find peaks
        peaks, _ = find_peaks(
            dataSelect, height=peakheight, prominence=peakprominence, width=peakwidth
        )

        # plot
        plt.plot(x, dataSelect)
        plt.plot(x[peaks], dataSelect[peaks], 'x')
        plt.yscale(plotscale)
        plt.xlabel(col_theta[i][1])
        plt.ylabel(col_counts[i][1])
        plt.title(col_counts[i][0])
        plt.show()

        # save peaks data in the list
        peaksOut = data[[col_theta[i], col_counts[i]]].loc[peaks]
        peaksOut.reset_index(drop=True, inplace=True)
        thePeakss.append(peaksOut)

        # save peaks data in the list
        dataCorr = np.vstack((x, dataSelect)).T
        # dataCorr = pd.DataFrame(data=dataCorrect, columns=column_headers)
        dataCorrect1.append(dataCorr)

    # convert list to dataframe
    thePeaks = pd.concat(thePeakss, axis=1)
    dataCorrected = np.concatenate(dataCorrect1, axis=1)
    dataCorrected = pd.DataFrame(dataCorrected, columns=data.columns)
    return thePeaks, dataCorrected


def XRD_background(
    data, peaks, cut_range=2, order=4, window_length=10
):
    Si_cut=True #foe now always disregard Si peak
    data_out = data.copy()
    headerlength = len(data.columns.get_level_values(1).unique())
    col_theta = data.columns.values[::2]
    col_counts = data.columns.values[1::2]
    peaks_theta = peaks.columns.values[::2]

    k = 0
    #variables for Si cut
    Si_start= 60
    Si_end = 70

    for i in range(0, len(col_theta)):
        cut_intensity = []

        two_theta = data[col_theta[i]]
        intensity = data[col_counts[i]]
        idx_range = np.where(two_theta >= 20 + cut_range)[0][0]

        # Cut data around peaks
        for j in range(len(intensity)):
            if data[col_theta[i]][j] in peaks[peaks_theta[i]].values:
                start_index = max(0, j - idx_range)
                end_index = min(len(data), j + idx_range)
                data_out[col_counts[i]][start_index:end_index] = (
                    np.nan
                )  # cut data intensity around peaks in data_out

        if Si_cut is True:
            idx_Si = np.where((two_theta >= Si_start) & (two_theta <= Si_end))[0]
            data_out[col_counts[i]][idx_Si] = np.nan

        cut_intensity = data_out[col_counts[i]]

        # Smooth the data for better peak detection
        smoothed_intensity = savgol_filter(
            intensity, window_length=window_length, polyorder=3
        )
        # Filter out NaN values (they exist because we cut the data) before fitting
        mask = ~np.isnan(cut_intensity)
        filtered_two_theta = two_theta[mask]
        filtered_intensity = intensity[mask]

        # Perform polynomial fitting with filtered data
        background_poly_coeffs = np.polyfit(
            filtered_two_theta, filtered_intensity, order
        )
        background = np.polyval(background_poly_coeffs, two_theta)

        # Subtract background
        corrected_intensity = smoothed_intensity - background

        data_out.insert(
            headerlength * (i + 1) + k,
            f'{data.columns.get_level_values(0).unique()[i]}',
            background,
            allow_duplicates=True,
        )
        data_out.rename(columns={'': 'Background'}, inplace=True)
        data_out.insert(
            headerlength * (i + 1) + k + 1,
            f'{data.columns.get_level_values(0).unique()[i]}',
            corrected_intensity,
            allow_duplicates=True,
        )
        data_out.rename(columns={'': 'Corrected Intensity'}, inplace=True)
        k = k + 2


        plt.figure()
        coord = data_out.columns.get_level_values(0).unique()[i]
        plt.plot(two_theta, intensity, label='Original Data')
        plt.plot(filtered_two_theta, filtered_intensity, label='filtered Data')
        plt.plot(
            two_theta,
            background,
            label='Background, order=' + str(order),
            linestyle='--',
        )
        plt.plot(two_theta, corrected_intensity, label='Corrected Data')
        plt.title(f'XRD data at {coord}')
        plt.legend()
        plt.show()
    # display(data_out)

    return data_out

def plot_XRD_shift_subplots(
    data,
    x,
    y_list,
    input_numbers, # nrows, ncols, shift!
    input_str, #datatype_x, datatype_y, title, material_guess

):
    """
    Plots XRD shift for multiple y-coordinates in subplots.

    Parameters:
    - data: DataFrame containing the data
    - datatype_x: str, type of data for x-axis
    - datatype_y: str, type of data for y-axis
    - x: list, list of x-coordinates
    - y_list: list of lists, each sublist contains y-coordinates for a subplot
    - shift: float, value by which to shift the y-axis data
    - title: str, title of the entire figure
    - ref_lines: array, reference lines to be plotted
    - nrows: int, number of rows of subplots
    - ncols: int, number of columns of subplots
    - figsize: tuple, size of the figure
    - save: bool, whether to save the figure as a file

    Returns:
    - plots the XRD data for multiple y-coordinates in subplots
    """
    figsize=(12, 10) #fix figure size
    nrows, ncols, shift= input_numbers
    datatype_x, datatype_y, title, material_guess = input_str

    with open(os.path.join('XRD', 'reflections', 'reflections.pkl'), 'rb') as file:
        ref_peaks_df = pickle.load(file)

    ref_peaks = ref_peaks_df[material_guess]
    ref_lines = ref_peaks['Peak 2theta'][ref_peaks['Peak 2theta'].notna()].values
    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=figsize)
    axes = axes.flatten()

    for idx, pos in enumerate(y_list):
        ax = axes[idx]
        for i in range(len(x)):
            # print('x =', x[i], 'y =', pos[i])
            x_data = get_data(
                data, datatype_x, x[i], pos[i], printinfo=False, drop_nan=False
            )
            y_data = get_data(
                data, datatype_y, x[i], pos[i], printinfo=False, drop_nan=False
            )
            lab = f'{x[i]:.1f},{pos[i]:.1f}'

            ax.plot(x_data, y_data + shift * i, label=lab)

        ax.set_title(f'Y = {pos[0]}')
        ax.legend(bbox_to_anchor=(1.01, 1), loc='upper left')

        if ref_lines is not None:
            for line in ref_lines:
                ax.axvline(x=line, linestyle='--', alpha=0.5, color='grey')

    axes[-1].plot(ref_peaks['2theta'], ref_peaks['I'], label=str(material_guess))
    # axes[-1].axvline(x=ref_lines.values, linestyle='--', alpha=0.5, color='grey')
    axes[-1].legend(bbox_to_anchor=(1.01, 1), loc='upper left')
    plt.suptitle(title)
    plt.tight_layout(rect=[0, 0, 1, 0.97])


    plt.savefig(f'{title}_XRD_shift_subplots.png', dpi=120, bbox_inches='tight')

    plt.show()


def plot_XRD_shift(
    data,
    datatype_x_y,
    shift,
    x_y, # f.e x_y = [[1, 2], [3, 4]]
    title=None
):  # x, y = list of points to plot]
    datatype_x, datatype_y = datatype_x_y
    x, y = x_y
    x_data = []
    y_data = []
    labels = []
    plt.figure(figsize=(12, 5))
    for i in range(len(x)):
        x_data.append(get_data(data, datatype_x, x[i], y[i], False, False))
        y_data.append(get_data(data, datatype_y, x[i], y[i], False, False))
        if x[0] == 'all' and y[0] == 'all':
            labels = data.columns.get_level_values(0).unique().values
        else:
            grid = MI_to_grid(data)
            xcoord, ycoord = closest_coord(grid, x[i], y[i])
            labels.append(f'{xcoord:.1f},{ycoord:.1f}')

        plt.plot(x_data[i], y_data[i] + shift * i, label=labels[i])
    plt.xlabel(datatype_x)
    plt.ylabel(datatype_y)
    plt.title(title)
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')

    plt.show()


def fit_two_related_peaks(x, y):
    # Initialize two Pseudo-Voigt models with prefixes to distinguish parameters
    model1 = PseudoVoigtModel(prefix='p1_')
    model2 = PseudoVoigtModel(prefix='p2_')

    # Estimate initial parameters for the first peak
    params = model1.guess(y, x=x)

    # Extract initial guesses
    amplitude1 = params['p1_amplitude'].value
    center1 = params['p1_center'].value
    sigma1 = params['p1_sigma'].value
    fraction1 = params['p1_fraction'].value

    # Set constraints for the second peak based on the provided relations
    # xpeak2 = 2 * np.arcsin((0.154439 / 0.1540562) * np.sin(center1 / 2))
    (360 / np.pi) * np.arcsin(
        (0.154439 / 0.1540562) * np.sin(center1 * np.pi / 360)
    )

    params.add(
        'p2_center',
        expr='(360/pi)* arcsin((0.154439 / 0.1540562) * sin(p1_center*pi /360))',
    )
    params.add('p2_amplitude', expr='0.5 * p1_amplitude')
    params.add('p2_sigma', expr='1 * p1_sigma')
    params.add('p2_fraction', expr='1 * p1_fraction')

    # Create a combined model by summing the two models
    combined_model = model1 + model2

    # Perform the fit
    fit_result = combined_model.fit(y, params, x=x)

    # Extract the fitted parameters for both peaks
    amplitude1 = fit_result.params['p1_amplitude'].value
    center1 = fit_result.params['p1_center'].value
    sigma1 = fit_result.params['p1_sigma'].value
    fraction1 = fit_result.params['p1_fraction'].value

    amplitude2 = fit_result.params['p2_amplitude'].value
    center2 = fit_result.params['p2_center'].value
    sigma2 = fit_result.params['p2_sigma'].value
    fraction2 = fit_result.params['p2_fraction'].value

    # Calculate FWHM for both peaks
    gamma1 = sigma1 / np.sqrt(2 * np.log(2))  # Convert sigma to gamma for Gaussian part
    fwhm1 = (1 - fraction1) * (2 * gamma1) + fraction1 * (2 * sigma1)

    gamma2 = sigma2 / np.sqrt(2 * np.log(2))
    fwhm2 = (1 - fraction2) * (2 * gamma2) + fraction2 * (2 * sigma2)

    return (
        fit_result,
        amplitude1,
        fwhm1,
        center1,
        fraction1,
        amplitude2,
        fwhm2,
        center2,
        fraction2,
    )


def fit_this_peak(data, peak_position, fit_range, withplots=True, printinfo=False):
    cut_range = fit_range
    peak_angle = peak_position

    dat_theta = data.iloc[:, data.columns.get_level_values(1) == '2θ (°)']
    dat_counts = data.iloc[:, data.columns.get_level_values(1) == 'Corrected Intensity']

    colors = plt.cm.jet(np.linspace(0, 1, len(dat_theta.columns)))

    plt.figure(figsize=(8, 6))

    df_fitted_peak = pd.DataFrame()

    for i in range(0, len(dat_theta.columns)):
        data_to_fit_x = dat_theta[dat_theta.columns[i]]
        data_to_fit_y = dat_counts[dat_counts.columns[i]]

        idx = np.where(
            (data_to_fit_x >= peak_angle - cut_range)
            & (data_to_fit_x <= peak_angle + cut_range)
        )[0]
        x_range = data_to_fit_x[idx].values
        y_range = data_to_fit_y[idx].values

        (
            fit_result,
            amplitude1,
            fwhm1,
            center1,
            fraction1,
            amplitude2,
            fwhm2,
            center2,
            fraction2,
        ) = fit_two_related_peaks(x_range, y_range)

        if printinfo is True:
            print(dat_theta.columns[i][0])
            print(
                f'Peak 1 - Amplitude: {amplitude1:.2f}, FWHM: {fwhm1:.2f}, '
                f'Center: {center1:.2f}, Fraction: {fraction1:.2f}'
            )
            print(
                f'Peak 2 - Amplitude: {amplitude2:.2f}, FWHM: {fwhm2:.2f}, '
                f'Center: {center2:.2f}, Fraction: {fraction2:.2f}'
            )

        if withplots is True:
            plt.plot(
                x_range,
                y_range,
                'o',
                color=colors[i],
                label=str(dat_theta.columns[i][0]),
            )
            plt.plot(x_range, fit_result.best_fit, '-', color=colors[i])
            plt.xlabel('2θ')
            plt.ylabel('Intensity')
            plt.title(' Fit with two related PseudoVoigts at ' + str(peak_angle) + '°')

        # store the information about the peak in a new dataframe

        peakData = np.vstack((center1, amplitude1, fwhm1, fraction1)).T
        peak_header = pd.MultiIndex.from_product(
            [[dat_theta.columns[i][0]], ['Center', 'Amplitude', 'FWHM', 'Fraction']],
            names=['Coordinate', 'Data type'],
        )
        df_peak_info = pd.DataFrame(data=peakData, columns=peak_header)
        fitData = np.vstack((x_range, y_range, fit_result.best_fit)).T
        fit_header = pd.MultiIndex.from_product(
            [[dat_theta.columns[i][0]], ['range 2θ', 'range Intensity', 'Fit']],
            names=['Coordinate', 'Data type'],
        )
        df_fit_info = pd.DataFrame(data=fitData, columns=fit_header)
        df_fitted_peak = pd.concat([df_fitted_peak, df_fit_info, df_peak_info], axis=1)

    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    # display(df_fitted_peak)
    return df_fitted_peak


def interactive_XRD_shift(
    data,
    datatype_x_y,
    shift,
    x_y,
    ref_peaks_df,
):
    """
    interactive shifted plot for assigning phases to XRD data, specify if you want
    different colors per each row or a rainbow colormap
    """

    datatype_x, datatype_y = datatype_x_y
    x, y = x_y
    colors='rows'


    fig = make_subplots(
        rows=2,
        cols=1,
        shared_xaxes=True,
        row_heights=[0.8, 0.2],  # Proportion of height for each plot
        vertical_spacing=0.02,  # Adjust this to reduce space between plots
    )

    if colors == 'rows':
        # Define a color palette with as many colors as there are unique values in y
        coords_colors = pd.DataFrame({'X': x, 'Y': y})
        unique_y_values = coords_colors['Y'].unique()

        color_palette = px.colors.qualitative.G10[: len(unique_y_values)]

        unique_x_values = coords_colors['X'].unique()
        color_dict = {}
        for i, color in enumerate(color_palette):
            # Generate lighter hues of the color for each x value
            base_color = mcolors.to_rgb(color)
            lighter_hues = [
                mcolors.to_hex(
                    (
                        base_color[0]
                        + (1 - base_color[0]) * (j / len(unique_x_values)),
                        base_color[1]
                        + (1 - base_color[1]) * (j / len(unique_x_values)),
                        base_color[2]
                        + (1 - base_color[2]) * (j / len(unique_x_values)),
                    )
                )
                for j in range(len(unique_x_values))
            ]
            color_dict[unique_y_values[i]] = lighter_hues
        coords_colors['Color'] = coords_colors.apply(
            lambda row: color_dict[row['Y']][list(unique_x_values).index(row['X'])],
            axis=1,
        )
        colors = coords_colors['Color'].values

    elif colors == 'rainbow':
        colormap = plt.get_cmap('turbo')  # You can choose any matplotlib colormap
        colors = [
            rgba_to_hex(colormap(i / len(x))) for i in range(len(x))
        ]  # Convert colors to hex

    x_data = []
    y_data = []
    # Store all y-data to find the global maximum
    all_y_data = []
    # Loop through and plot the XRD spectra with a vertical shift in the top plot
    for i in range(len(x)):
        x_data = get_data(data, datatype_x, x[i], y[i], False, False)
        y_data = get_data(data, datatype_y, x[i], y[i], False, False)
        shifted_y_data = y_data - shift * i

        all_y_data.extend(
            shifted_y_data
        )  # Collect y-data with shift for max computation

        fig.add_trace(
            go.Scatter(
                x=x_data,
                y=shifted_y_data,
                mode='lines',
                line=dict(color=colors[i]),
                name=f'{i+1}: {x[i]}, {y[i]}',
            ),
            row=1,
            col=1,
        )

    # Compute the global maximum y-value, considering shifts
    global_min_y = min(all_y_data)

    # Create traces for each reference material (hidden initially)
    ref_traces = []
    buttons = []

    for ref_material, ref_df in ref_peaks_df.items():
        # Reference spectrum plotted in the bottom plot
        ref_trace = go.Scatter(
            x=ref_df['2theta'],
            y=ref_df['I'],
            mode='lines',
            name=f'{ref_material} Reference',
            visible=False,
        )

        # Create vertical peak lines for top plot (raw data plot)
        peak_lines = go.Scatter(
            x=[
                value for peak in ref_df['Peak 2theta'] for value in [peak, peak, None]
            ],  # x: peak, peak, None to break the line
            y=[global_min_y, 1000 * 1.1, None]
            * len(
                ref_df['Peak 2theta']
            ),  # y: 0 -> global_max_y for each line, with None to break lines
            mode='lines',
            line=dict(color='grey', dash='dot'),
            showlegend=False,
            visible=False,
        )

        # Append traces for each reference spectrum and its peaks
        ref_traces.append(ref_trace)
        ref_traces.append(peak_lines)

        # Create a button for each reference
        buttons.append(
            dict(
                label=ref_material,
                method='update',
                args=[
                    {
                        'visible': [True] * len(x) + [False] * len(ref_traces)
                    },  # Show all raw spectra, hide refs by default
                    {'title': f'with {ref_material} Reference'},
                ],
            )
        )

    # Add reference traces to figure (initially hidden)
    for trace in ref_traces:
        # Ensure trace.name is not None before checking 'Reference' in name
        fig.add_trace(
            trace, row=2 if trace.name and 'Reference' in trace.name else 1, col=1
        )

    # Update buttons to control the visibility of one reference at a time
    for i, button in enumerate(buttons):
        # Make the selected reference spectrum visible in the bottom plot and its peaks
        # visible in the top plot
        button['args'][0]['visible'][len(x) :] = [False] * len(
            ref_traces
        )  # Hide all refs initially
        button['args'][0]['visible'][len(x) + 2 * i : len(x) + 2 * i + 2] = [
            True,
            True,
        ]  # Show selected ref and peaks

    # Add the dropdown menu to switch between reference spectra
    fig.update_layout(
        updatemenus=[
            {
                'buttons': buttons,
                'direction': 'down',
                'showactive': True,
                'x': 1.05,
                'xanchor': 'left',
                'y': 1.1,
                'yanchor': 'top',
            }
        ],
        template='plotly_white',  # Choose a template (e.g., 'plotly_dark')
        #title=title,
        height=600,  # Adjust the height of the figure (e.g., 700)
        width=900,  # Adjust the width of the figure (e.g., 900)
        legend=dict(x=1.05, y=1),
        xaxis2_title=datatype_x,
        yaxis_title=datatype_y,
    )

    fig.show()
    return fig

def rgba_to_hex(rgba):
    """Convert an RGBA tuple to a hex color string."""
    r, g, b, a = (int(c * 255) for c in rgba)
    return f'#{r:02x}{g:02x}{b:02x}'

def assign_phases_labels(data):
    """Function to assign phases to specific points in a dataset.
    Returns:
        phase_info (dict): Dictionary where the key is the phase and the value is a
            list of 'unknown', 'amorphous', or the phase name corresponding to the
            presence of that phase at each coordinate.
    """
    coords = data.columns.get_level_values(0).unique()

    phase_info = {}  # Dictionary to store phase information for each point
    num_coords = len(coords)

    # Initialize the presence array with 'unknown'
    phase_present = ['unknown'] * num_coords

    # Ask user for the main phase name
    main_phase = input(
        "What is the main phase present? (or type 'exit' to finish): "
    ).strip()
    if main_phase.lower() == 'exit':
        phase_info['Phase'] = phase_present
        return phase_info  # Return the dictionary with 'unknown' if user exits

    # Assign the main phase to all points initially
    phase_present = [main_phase] * num_coords

    while True:
        # Ask if there is any other phase
        other_phase_response = (
            input('Is there any other phase present? (yes/no): ').strip().lower()
        )
        if other_phase_response == 'no':
            break

        # Ask user for the other phase name
        other_phase = input(
            "What is the other phase name? (or type 'exit' to finish): "
        ).strip()
        if other_phase.lower() == 'exit':
            break

        # Display available points
        print('\nAvailable points (coordinates):')
        for i, coord in enumerate(coords):
            print(f'{i + 1}: {coord}')

        # Ask which points should be set to the other phase
        selected_points = input(
            f"\nWhich points should be set to '{other_phase}'? "
            '(Enter numbers separated by commas): '
        ).strip()
        selected_indices = [
            int(idx.strip()) - 1
            for idx in selected_points.split(',')
            if idx.strip().isdigit()
        ]
        for idx in selected_indices:
            if 0 <= idx < num_coords:
                phase_present[idx] = other_phase

    # Store this phase's information in the dictionary
    phase_info['Phase'] = phase_present
    print(f"\nPhase '{main_phase}' assigned to the remaining points.")
    print(phase_info)

    return phase_info


def assign_phases_numbers(data):
    # obsolete, use phase labels instead
    """Function to assign phases to specific points in a dataset.  coords
    (list of tuples): List of coordinates available for selection.
    Returns:
        phase_info (dict): Dictionary where the key is the phase and the value is a
            list of 'yes'/'no' corresponding to the presence of that phase at each
            coordinate.
    """
    coords = data.columns.get_level_values(0).unique()

    phase_info = {}  # Dictionary to store phase information for each point
    num_coords = len(coords)

    # Ask user for the phase name
    phase = input("What is the phase name? (or type 'exit' to finish): ").strip()
    if phase.lower() == 'exit':
        return phase_info  # Return empty dictionary if user exits

    # Determine if phase is present in most or few points
    presence_type = (
        input(
            'Is the phase present in most points or few points? '
            "(type 'most' or 'few'): "
        )
        .strip()
        .lower()
    )

    # Initialize the presence array based on user input
    if presence_type == 'most':
        phase_present = [1] * num_coords  # Start with all points as 'yes'
    elif presence_type == 'few':
        phase_present = [0] * num_coords  # Start with all points as 'no'
    else:
        print("Invalid input. Please enter 'most' or 'few'.")

    # Display available points
    print('\nAvailable points (coordinates):')
    for i, coord in enumerate(coords):
        print(f'{i + 1}: {coord}')

    # Ask which points should be changed
    if presence_type == 'most':
        selected_points = input(
            "\nWhich points should be set to 'no'? "
            '(Enter numbers separated by commas): '
        ).strip()
        selected_indices = [
            int(idx.strip()) - 1
            for idx in selected_points.split(',')
            if idx.strip().isdigit()
        ]
        for idx in selected_indices:
            if 0 <= idx < num_coords:
                phase_present[idx] = 0
    else:  # presence_type == 'few'
        selected_points = input(
            "\nWhich points should be set to 'yes'? "
            '(Enter numbers separated by commas): '
        ).strip()
        selected_indices = [
            int(idx.strip()) - 1
            for idx in selected_points.split(',')
            if idx.strip().isdigit()
        ]
        for idx in selected_indices:
            if 0 <= idx < num_coords:
                phase_present[idx] = 1

    # Store this phase's information in the dictionary
    phase_info[phase] = phase_present
    print(f"\nPhase '{phase}' assigned to the selected points.")
    print(phase_info)

    return phase_info


##########################
# Functions related to UPS analysis
##########################

#redo this function with more time and care!!!! Too many assumptions and wierd fits
# def UPS_fit(
#     data,
#     startvalue,
#     guess_slope=2000,
#     slope_change=1.5,
#     window=1,
#     background_end=None,
#     fit_background = True,
#     plotscale = 'linear'
#     ):
#     '''New: actually never used to fit semiconductor data.
#     Fit UPS data using dataframe input from read_UPS,
#     the data is fitted by finding background and valence bands based on slope,
#     they are both fitted
#     Output: dataframe with the background and fit intersection point,
#     as well as the backgrounds intersect with x'''
#     column_headers = data.columns.values
#     col_BE = column_headers[::2]
#     col_counts = column_headers[1::2]
#     mod = LinearModel(prefix='reg_')


#     UPS_outframe = pd.DataFrame()
#     for i in range(0,len(col_BE)):
#         x = data[col_BE[i]]
#         y = data[col_counts[i]]

#         #reversing the data
#         x_reversed = x[::-1].values
#         y_reversed = y[::-1].values
#         #y_reversed = savgol_filter(y[::-1].values, 2, 1)

#         #select based on input start value
#         xselect = x_reversed[list(range(startvalue,len(x)))]
#         yselect = y_reversed[list(range(startvalue,len(y)))]

#         #finding the background onset from slope
#         k = 1
#         slope = None
#         while k < len(yselect)-1:
#             slope = yselect[k]-yselect[k-window]
#             if slope < guess_slope:
#                 k = k + 1
#             else:
#                 bkg_start = k
#                 bkg_slope = slope
#                 k = len(yselect)-1

#         #finding the background end from onset point and slope
#         xselect1 = xselect[list(range(bkg_start,len(xselect)))]
#         yselect1 = yselect[list(range(bkg_start,len(yselect)))]

#         k = 25
#         v= 100
#         slope = None
#         while k < v:
#             slope = yselect1[k]-yselect1[k-window]
#             if slope < bkg_slope*slope_change:
#                 k = k + 1
#             else:
#                 bkg_end = k
#                 k = 101

#         if background_end:
#             bkg_end = np.where(xselect1 == background_end)[0][0]

#         #fitting the background with a linear model
#         x_bkg = xselect1[list(range(0,bkg_end))]
#         y_bkg = yselect1[list(range(0,bkg_end))]
#         out_bkg = mod.fit(y_bkg, x = x_bkg)
#         bkg_x_intercept = (
#             -out_bkg.params['reg_intercept'].value/out_bkg.params['reg_slope'].value
#             )

#         if fit_background:
#             #finding the valence onset

#             xselect_valence = x_reversed[
#                 list(range(bkg_end+bkg_start+startvalue,len(x)))
#                 ]
#             yselect_valence = y_reversed[
#                 list(range(bkg_end+bkg_start+startvalue,len(y)))
#                 ]

#             k = 5
#             slope = None
#             while k < len(yselect_valence)-1:
#                 slope = yselect_valence[k]-yselect_valence[k-1]
#                 if slope < bkg_slope*4:
#                     k = k + 1
#                 else:
#                     valence_start = k
#                     valence_slope = slope
#                     k = len(yselect_valence)-1

#             #finding valence end
#             xselect_valence1 = xselect_valence[
#                 list(range(valence_start,len(xselect_valence)))
#                 ]
#             yselect_valence1 = yselect_valence[
#                 list(range(valence_start,len(yselect_valence)))
#                 ]

#             k = 15
#             slope = None
#             while k < len(yselect_valence1)-1:
#                 slope = yselect_valence1[k]-yselect_valence1[k-1]
#                 if slope < valence_slope*1.2:
#                     k = k + 1
#                     valence_end = len(yselect_valence1)-2
#                 else:
#                     valence_end = k -1
#                     k = len(yselect_valence1)
#             try:
#                 x_valence = xselect_valence1[list(range(0,valence_end))]
#                 y_valence = yselect_valence1[list(range(0,valence_end))]
#                 out_valence = mod.fit(y_valence, x = x_valence)
#                 fits_intercept = ((
#                     out_valence.params['reg_intercept'].value
#                     - out_bkg.params['reg_intercept'].value
#                     )/(
#                     out_bkg.params['reg_slope'].value
#                     - out_valence.params['reg_slope'].value)
#                     )

#             except(AttributeError, TypeError) as e:
#                 print('No valence band found'+ e)
#                 out_valence = None



#             plt.plot(x,y, label = 'data')
#             try:
#                 plt.plot(x_valence,out_valence.best_fit, label = 'valence fit')
#                 plt.plot(
#                     xselect_valence[valence_start],yselect_valence[valence_start],
#                     'o',
#                     label = 'valence_start',
#                     )
#                 plt.plot(
#                     xselect_valence1[valence_end],yselect_valence1[valence_end],
#                     'o',
#                     label = 'valence_end',
#                     )
#             except(AttributeError, TypeError) as e:
#                 print(f"An error occurred: {e}")
#                 pass
#             plt.plot(xselect[bkg_start], yselect[bkg_start], 'o',label = 'bkg_start')
#             plt.plot(xselect1[bkg_end], yselect1[bkg_end], 'o',label = 'bkg_end')
#             plt.plot(x_bkg, out_bkg.best_fit, '--',label='background fit')
#             plt.xlabel(col_BE[i][1])
#             plt.ylabel(col_counts[i][1])
#             plt.title(col_counts[i][0])
#             plt.yscale(plotscale)
#             plt.xlim(-1, 4)
#             # plt.ylim(-1e4,2e6)
#             plt.legend()
#             plt.show()

#             try:
#                 print("bkg_x_intercept:\n",bkg_x_intercept)
#                 print("fits_intercept:\n",fits_intercept)
#             except(AttributeError, TypeError) as e:
#                 print(f"An error occurred: {e}")
#                 fits_intercept = None*len(bkg_x_intercept)

#             intercepts = np.vstack((bkg_x_intercept,fits_intercept)).T
#             UPS_header = pd.MultiIndex.from_product(
#                 [[col_BE[i][0]],
#                 ['bkg_x_intercept','fits_intercept']],
#                 names=['Coordinate','Data type'],
#                 )
#             UPS_output = pd.DataFrame(data=intercepts, columns=UPS_header)


#         else:
#             plt.plot(x,y, label = 'data')
#             plt.plot(x_bkg, out_bkg.best_fit, label='valence fit')
#             plt.xlabel(col_BE[i][1])
#             plt.ylabel(col_counts[i][1])
#             plt.title(col_counts[i][0])
#             plt.yscale(plotscale)
#             plt.xlim(-1, 4)
#             # plt.ylim(-1e4,2e6)
#             plt.legend()
#             plt.show()

#             print("bkg_x_intercept:\n",bkg_x_intercept)

#         UPS_header = pd.MultiIndex.from_product(
#             [[col_BE[i][0]],
#             ['bkg_x_intercept']],
#             names=['Coordinate','Data type'],
#             )
#         UPS_output = pd.DataFrame(
#             data=np.array([bkg_x_intercept])
#             , columns=UPS_header,
#             )

#         UPS_outframe = pd.concat([UPS_outframe, UPS_output], axis = 1)
#     return UPS_outframe

def adjust_BE_values(data):
    """
    Adjusts the BE values in the DataFrame.
    If a BE value is higher than 10^4, it divides it by 10^5.

    Parameters:
    data (pd.DataFrame): The input DataFrame containing BE values.

    Returns:
    pd.DataFrame: The DataFrame with adjusted BE values.
    """
    # Check if the 'BE' column exists in the DataFrame at level 1
    if 'BE (eV)' in data.columns.get_level_values(1):
        # Apply the adjustment to the 'BE' column
        data.loc[:, (slice(None), 'BE (eV)')] = data.loc[
            :,
            (slice(None), 'BE (eV)')
            ].map(lambda x: x / 10**5 if x > 10**4 else x)
    else:
        print("The DataFrame does not contain a 'BE (eV)' column at level 1.")

    return data


##########################
# Functions related to REELS analysis
##########################


def REELS_fit(data, plotscale='linear'):
    """Fit REELS data using dataframe input from read_REELS, the data is fitted by
    finding the the 0 energy loss peak position, and the onset of the energy loss curve
    Output: dataframe with the onset and peak positions, as well as the band gap and the
      raw data."""
    column_headers = data.columns.values
    column_headers = data.columns.values
    col_BE = column_headers[::2]
    col_counts = column_headers[1::2]
    LinearModel(prefix='reg_')

    BGS = np.array([])
    REELS_outframe = pd.DataFrame()
    REELS_data = pd.DataFrame()
    for i in range(0, len(col_BE)):
        x = data[col_BE[i]]
        y = data[col_counts[i]]

        # finds the 0 energy loss peak
        peaks, _ = find_peaks(y, prominence=1000000)

        # selects the data after the peak
        xselect1 = x[list(range(0, peaks[0]))].values
        yselect1 = y[list(range(0, peaks[0]))].values

        # formating
        yselect1 = yselect1[::-1]
        xselect1 = xselect1[::-1]
        peak = x[peaks].values[0]

        # locates the onset of the energy loss.
        j = False
        k = 0
        v=100
        slope_cutof = 1400
        while j is not True:
            slope = yselect1[k + 1] - yselect1[k]
            if slope < slope_cutof:
                k = k + 1
            else:
                idx2 = k
                j = True
            if k == v:
                j = True

        onset = xselect1[idx2]
        # finds the onset from the peak location and onset
        BG = onset - peak

        plt.plot(x, y, label='data')
        plt.plot(x[peaks], y[peaks], 'o', label='elastic scattering peak')
        plt.plot(xselect1[idx2], yselect1[idx2], 'o', label='onset of energy loss')
        plt.xlabel(col_BE[i][1])
        plt.ylabel(col_counts[i][1])
        plt.title(col_counts[i][0])
        plt.yscale(plotscale)
        plt.xlim(20, 4)
        plt.legend()
        plt.show()

        # print output peak positions, intensity, and FWHM
        print(f"Onset:\n{onset}\nPeak:\n{peak}\nBand gap:\n{BG}")

        # constructs output dataframe of band gap, onset and Peak values.
        BGS = np.vstack((BG, xselect1[idx2], x[peaks])).T
        REELS_header = pd.MultiIndex.from_product(
            [[col_BE[i][0]], ['Band gap', 'Onset BE', 'Peak BE']],
            names=['Coordinate', 'Data type'],
        )
        REELS_bandgaps = pd.DataFrame(data=BGS, columns=REELS_header)

        # construct output for the raw data
        raw = np.vstack((x, y)).T
        raw_header = pd.MultiIndex.from_product(
            [[col_BE[i][0]], ['BE (ev)', 'Intensity (counts)']],
            names=['Coordinate', 'Data type'],
        )
        REELS_data = pd.DataFrame(data=raw, columns=raw_header)

        # assembles the final output
        REELS_outframe = pd.concat([REELS_outframe, REELS_data, REELS_bandgaps], axis=1)
    return REELS_outframe


##########################
# Functions related to plotting
##########################


def plot_grid(coords, grid):
    """Plot a set of real measurement points on a custom grid defined with the
    "measurement_grid" function. The corrected grid locations are shown."""
    corrected_grid = coords_to_grid(coords, grid)
    plt.scatter(grid.iloc[:, 0], grid.iloc[:, 1], color='black', s=80)
    plt.scatter(coords.iloc[:, 0], coords.iloc[:, 1], color='green', s=20)
    plt.scatter(corrected_grid.iloc[:, 0], corrected_grid.iloc[:, 1], color='red', s=20)
    plt.legend(['Defined grid', 'Measured', 'Corrected'])


def plot_data(
    data,
    datatype_x_y,
    x_y=['all','all'],
    select=[None,None, False],
    #scatter_plot=False,
    title='auto',
):
    """Creates a XY plot/scatter plot based on datatype from a dataframe"""

    # x and y to list if only 1 value specified
    datatype_x, datatype_y = datatype_x_y
    x , y = x_y
    datatype_select, datatype_select_value, scatter_plot = select
    plotscale='linear',

    x = [x] if not isinstance(x, list) else x
    y = [y] if not isinstance(y, list) else y
    x_data = []
    y_data = []
    labels = []
    # extracts the specified data point by point
    for i in range(len(x)):
        x_data.append(get_data(data, datatype_x, x[i], y[i], False, False))
        y_data.append(get_data(data, datatype_y, x[i], y[i], False, False))
        if x[0] == 'all' and y[0] == 'all':
            labels = data.columns.get_level_values(0).unique().values
        else:
            grid = MI_to_grid(data)
            xcoord, ycoord = closest_coord(grid, x[i], y[i])
            labels.append(f'{xcoord:.1f},{ycoord:.1f}')

    colors = plt.cm.jet(
        np.linspace(0, 1, len(labels))
    )  # data.columns.get_level_values(0).unique().values

    # formating
    if len(labels) == 1:
        labels = labels[0]
    if x[0] == 'all' and y[0] == 'all':
        x_data, y_data = x_data[0], y_data[0]
    else:
        x_data, y_data = np.transpose(x_data), np.transpose(y_data)

    # if datatype with multiple values per point is plotted only selects one value,
    # based on the datatype_select, datatype_select_value.
    if datatype_select is not None:
        y_data = y_data.iloc[
            data.index[
                data[
                    data.iloc[
                        :, data.columns.get_level_values(1) == datatype_select
                    ].columns[0]
                ]
                == datatype_select_value
            ]
        ]
        x_data_coords = x_data.columns.get_level_values(0)
        y_data_coords = y_data.columns.get_level_values(0)
        data_coords = [j for j in x_data_coords if j not in y_data_coords]
        x_data.drop(data_coords, level=0, axis=1, inplace=True)
        x_data = x_data.values[0]
        y_data = y_data.values[0]
        labels = datatype_select + ': ' + str(round(datatype_select_value, 2))

    # plots scatter plot if scatter_plot is not false, else line plot
    # Determine the data to use based on the condition
    x_values = x_data.values.T if x[0] == 'all' else x_data.T
    y_values = y_data.values.T if y[0] == 'all' else y_data.T
    #now plot check that this change works
    for idx, (x_val, y_val) in enumerate(zip(x_values, y_values)):
        if scatter_plot:
            plt.plot(x_val, y_val, 'o', color=colors[idx], label=labels[idx])
        else:
            plt.plot(x_val, y_val, color=colors[idx], label=labels[idx])

    plt.xlabel(datatype_x)
    plt.ylabel(datatype_y)
    plt.yscale(plotscale)
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    if title == 'auto':
        plt.title(f'{datatype_y} over {datatype_x}')
    else:
        plt.title(title)


def new_heatmap(
    datatype,
    data=None,
    exclude=None,
    cbar=None,
    index=None,
):
    """
    plot heatmaps with interpolated background, like in Nomad, if savepath ends with
    .png, it will save as png, if it ends with .html, it will save as html (interactive)
    """


    if data is not None:
        if exclude is not None:
            for point in exclude:
                data = data.drop(
                    data.iloc[:, data.columns.get_level_values(0) == point], axis=1
                )
        xy = MI_to_grid(data).drop_duplicates(ignore_index=True)
        x = xy['x'].values
        y = xy['y'].values
        if index is None:
            z = (
                data.iloc[:, data.columns.get_level_values(1) == datatype]
                .dropna()
                .values.flatten()
            )
        else:
            z = (
                data.iloc[index, data.columns.get_level_values(1) == datatype]
                .dropna()
                .values.flatten()
            )

    xi = np.linspace(min(x), max(x), 100)
    yi = np.linspace(min(y), max(y), 100)
    xi, yi = np.meshgrid(xi, yi)
    zi = griddata((x, y), z, (xi, yi), method='linear')

    scatter = go.Scatter(
        x=x,
        y=y,
        mode='markers',
        marker=dict(
            size=15,
            color=z,  # Set color to thickness values
            colorscale='Viridis',  # Choose a colorscale
            showscale=False,  # Hide the colorbar for the scatter plot
            line=dict(
                width=2,  # Set the width of the border
                color='DarkSlateGrey',  # Set the color of the border
            ),
        ),
    )
    if datatype == 'Layer 1 Thickness (nm)':
        cbar_title = 'Thickness (nm)'
    elif datatype.startswith('Layer 1 ') and datatype.endswith(' Atomic %'):
        element = datatype.split()[2]
        cbar_title = f'{element} Atomic %'
        #hope this works (simplifies the code but not tested)
    else:
        cbar_title = datatype
    if cbar is not None:
        cbar_title = cbar

    heatmap = go.Heatmap(
        x=xi[0],
        y=yi[:, 0],
        z=zi,
        colorscale='Viridis',
        colorbar=dict(title=cbar_title),
        # zmin = 10, zmax = 60
    )

    fig = go.Figure(data=[heatmap, scatter])


    title = datatype

    fig.update_layout(
        title=title,
        xaxis_title='X Position (mm)',
        yaxis_title='Y Position (mm)',
        template='plotly_white',
        autosize=False,
        width=600,
        height=500,
    )


    fig.show()


def plot_scatter_colormap(
    data,
    datatype_x_y_z,
    x_y=['all','all'],
    select=[None, None,None], #datatype_select, datatype_select_value, colormap_label
    limits=[None,None],
):
    """Creates a XY plot/scatter plot based on datatype"""
    datatype_x, datatype_y, datatype_z = datatype_x_y_z
    x, y = x_y
    min_limit, max_limit = limits
    plotscale='linear'
    title='auto'
    datatype_select, datatype_select_value, colormap_label = select
    # x and y to list if only 1 value specified
    #if not isinstance(x, list):
    #    x = [x]
    #if not isinstance(y, list):
    #    y = [y]
    #x_data = []
    #y_data = []
    #z_data = []

    # new copilot code. Test!!!!
    x = [x] if not isinstance(x, list) else x
    y = [y] if not isinstance(y, list) else y

    # Extract data
    x_data = [get_data(data, datatype_x, xi, yi, False, False) for xi, yi in zip(x, y)]
    y_data = [get_data(data, datatype_y, xi, yi, False, False) for xi, yi in zip(x, y)]
    z_data = [get_data(data, datatype_z, xi, yi, False, False) for xi, yi in zip(x, y)]


    labels = []
    # extracts the specified data point by point
    for i in range(len(x)):
        #x_data.append(get_data(data, datatype_x, x[i], y[i], False, False))
        #y_data.append(get_data(data, datatype_y, x[i], y[i], False, False))
        #z_data.append(get_data(data, datatype_z, x[i], y[i], False, False))
        if x[0] == 'all' and y[0] == 'all':
            labels = data.columns.get_level_values(0).unique().values

        else:
            grid = MI_to_grid(data)
            xcoord, ycoord = closest_coord(grid, x[i], y[i])
            labels.append(f'{xcoord:.1f},{ycoord:.1f}')

    # formating
    labels = labels[0] if len(labels) == 1 else labels
    if x[0] == 'all' and y[0] == 'all':
        x_data, y_data, z_data = x_data[0], y_data[0], z_data[0]
    else:
        x_data, y_data, z_data = map(np.transpose, [x_data, y_data, z_data])

    # if datatype with multiple values per point is plotted only selects one value,
    # based on the datatype_select, datatype_select_value.
    if datatype_select is not None:
        z_data = z_data.iloc[
            data.index[
                data[
                    data.iloc[
                        :, data.columns.get_level_values(1) == datatype_select
                    ].columns[0]
                ]
                == datatype_select_value
            ]
        ]
        x_data_coords = x_data.columns.get_level_values(0)
        z_data_coords = z_data.columns.get_level_values(0)
        data_coords = [j for j in x_data_coords if j not in z_data_coords]
        x_data.drop(data_coords, level=0, axis=1, inplace=True)
        y_data.drop(data_coords, level=0, axis=1, inplace=True)
        x_data, y_data, z_data = x_data.values[0], y_data.values[0], z_data.values[0]

    # removes data points above the max limit of the z data
    if max_limit is not None:
        x_data = x_data[z_data < max_limit]
        y_data = y_data[z_data < max_limit]
        z_data = z_data[z_data < max_limit]

    # removes data points below the min limit of the z data
    if min_limit is not None:
        x_data = x_data[z_data > min_limit]
        y_data = y_data[z_data > min_limit]
        z_data = z_data[z_data > min_limit]

    # Defines the color from the z_data
    colors = z_data
    plt.scatter(x_data, y_data, c=colors, cmap='viridis')
    plt.colorbar().set_label(colormap_label)
    plt.xlabel(datatype_x)
    plt.ylabel(datatype_y)
    plt.yscale(plotscale)
    if title == 'auto':
        plt.title(f'{datatype_y} over {datatype_x}')
    else:
        plt.title(title)

def ternary_plot(df, element_list, datatype, title):
    """make a ternary plot of the data in df, with el1, el2, el3 as the corners, and
    colorscale based on datatype"""
    el1=element_list[0]
    el2=element_list[1]
    el3=element_list[2]

    A = f'Layer 1 {el1} Atomic %'
    B = f'Layer 1 {el2} Atomic %'
    C = f'Layer 1 {el3} Atomic %'

    A_percent = get_data(df, A).loc[0].values.flatten()
    B_percent = get_data(df, B).loc[0].values.flatten()
    C_percent = get_data(df, C).loc[0].values.flatten()
    intensity = get_data(df, datatype).loc[0]
    X, Y = extract_coordinates(df)

    custom_data = list(zip(X, Y, intensity))

    fig = go.Figure()
    fig.add_trace(
        go.Scatterternary(
            {
                'mode': 'markers',
                'a': A_percent,  # el1 percentages
                'b': B_percent,  # el2 percentages
                'c': C_percent,  # el3 percentages
                'marker': {
                    'symbol': 100,
                    'size': 8,
                    'color': intensity,  # Use intensity for marker color
                    'colorscale': 'Turbo',  # Choose a colorscale
                    'colorbar': {'title': datatype},  # Add a colorbar
                    'line': {'width': 2},
                },
                # 'text': coordinates,
                'customdata': custom_data,
                'hovertemplate': (
                    f'{el1}: %{{a:.1f}}%<br>{el2}: %{{b:.1f}}%<br>{el3}: %{{c:.1f}}%'
                    f'<br>{datatype}: %{{marker.color:.1f}}'
                    f'<br>Coordinates: (%{{customdata[0]:.1f}}, %{{customdata[1]:.1f}})'
                ),
                'showlegend': False,
            }
        )
    )

    # Update layout
    fig.update_layout(
        {
            'ternary': {
                'sum': 100,
                'aaxis': {
                    'title': f'{el1} %',
                    'min': 0,
                    'linewidth': 2,
                    'ticks': 'outside',
                },
                'baxis': {
                    'title': f'{el2} %',
                    'min': 0,
                    'linewidth': 2,
                    'ticks': 'outside',
                },
                'caxis': {
                    'title': f'{el3} %',
                    'min': 0,
                    'linewidth': 2,
                    'ticks': 'outside',
                },
            },
            'title': title,
        },
        width=800,
        height=600,
    )

    # Show the plot
    fig.show()




def ternary_discrete(
    df,
    element_list,
    labels,
    title,
    include_id=True,
):
    """
    Create a ternary plot with discrete colors for string intensities and different
    marker shapes for phases.

    Parameters:
    df (pd.DataFrame): DataFrame containing the data.
    el1 (str): Label for the first element.
    el2 (str): Label for the second element.
    el3 (str): Label for the third element.
    intensity_label (str): Label for the intensity values.
    shape_label (str): Label for the phase values.
    title (str): Title of the plot.
    """
    intensity_label, shape_label = labels
    el1 = element_list[0]
    el2 = element_list[1]
    el3 = element_list[2]

    # Extract element percentages, intensity, and phase
    A = f'Layer 1 {el1} Atomic %'
    B = f'Layer 1 {el2} Atomic %'
    C = f'Layer 1 {el3} Atomic %'

    A_percent = get_data(df, A).loc[0].values.flatten()
    B_percent = get_data(df, B).loc[0].values.flatten()
    C_percent = get_data(df, C).loc[0].values.flatten()
    intensity = get_data(df, intensity_label).loc[0]
    phase = get_data(df, shape_label).loc[0]
    if include_id is True:
        sample_id = get_data(df, 'Sample ID').loc[0]
    elif include_id is False:
        sample_id = ['unknown sample'] * len(intensity)
    # coords = MI_to_grid(df).values
    X, Y = extract_coordinates(df)  # problems if more than one point with same coords

    # Create a color mapping for unique intensity values

    unique_intensities = list(set(intensity))
    color_map = {val: i for i, val in enumerate(unique_intensities)}
    colors = [color_map[val] for val in intensity]

    # Create a marker shape mapping for unique phase values
    unique_phases = list(set(phase))
    marker_shapes = [
        'circle',
        'square',
        'diamond',
        'cross',
        'x',
        'triangle-up',
        'triangle-down',
        'triangle-left',
        'triangle-right',
    ]
    shape_map = {
        val: marker_shapes[i % len(marker_shapes)]
        for i, val in enumerate(unique_phases)
    }
    shapes = [shape_map[val] for val in phase]

    # custom_data = list(zip(X,Y, intensity, phase))
    custom_data = list(zip(X, Y, intensity, phase, sample_id))

    # Create the ternary plot with custom hover text, colored markers, and different
    # shapes
    fig = go.Figure(
        go.Scatterternary(
            {
                'mode': 'markers',
                'a': A_percent,  # el1 percentages
                'b': B_percent,  # el2 percentages
                'c': C_percent,  # el3 percentages
                'marker': {
                    'size': 8,
                    'color': colors,  # Use mapped colors for marker color
                    'symbol': shapes,  # Use mapped shapes for marker shape
                    'colorscale': 'Turbo',  # Choose a colorscale
                    'colorbar': {
                        'title': intensity_label,
                        'tickvals': list(
                            color_map.values()
                        ),  # Set tick values to the mapped color indices
                        'ticktext': unique_intensities,  # Set tick text to the unique
                        # intensity values
                    },
                    'line': {'width': 1},
                },
                'customdata': custom_data,  # Store the x-coordinates in custom data
                # 'text': coords,#df.columns.get_level_values(0).unique(),  # Labels
                # for the points
                # 'hovertemplate': f'{el1}: %{{a:.1f}}%<br>{el2}: %{{b:.1f}}%<br>{el3}:
                # %{{c:.1f}}%<br>Coordinates: %{{text}}<br>{intensity_label}:
                # %{{marker.color}}%',  # Custom hover text format
                'hovertemplate': (
                    f'{el1}: %{{a:.1f}}%<br>{el2}: %{{b:.1f}}%<br>{el3}: %{{c:.1f}}%'
                    f'<br>Coordinates: (%{{customdata[0]:.1f}}, %{{customdata[1]:.1f}})'
                    f'<br>{intensity_label}: %{{customdata[2]}}'
                    f'<br>{shape_label}: %{{customdata[3]}}'
                    f'<br>Sample: %{{customdata[4]:.0f}}'
                ),
                'name': 'Data Points',
                'showlegend': False,
            }
        )
    )

    # Add dummy traces for the legend
    for phase_name, shape in shape_map.items():
        fig.add_trace(
            go.Scatterternary(
                {
                    'mode': 'markers',
                    'a': [None],  # Dummy data
                    'b': [None],  # Dummy data
                    'c': [None],  # Dummy data
                    'marker': {'size': 8, 'symbol': shape, 'color': 'black'},
                    'name': phase_name,
                    'showlegend': True,
                }
            )
        )

    # Update layout
    fig.update_layout(
        {
            'ternary': {
                'sum': 100,
                'aaxis': {
                    'title': f'{el1} %',
                    'min': 0,
                    'linewidth': 2,
                    'ticks': 'outside',
                },
                'baxis': {
                    'title': f'{el2} %',
                    'min': 0,
                    'linewidth': 2,
                    'ticks': 'outside',
                },
                'caxis': {
                    'title': f'{el3} %',
                    'min': 0,
                    'linewidth': 2,
                    'ticks': 'outside',
                },
            },
            'title': title,
            'legend': {'x': -0.1, 'y': 1},  # Position the legend on the left
        },
        width=800,
        height=600,
    )

    # Show the plot
    fig.show()


##########################
# Functions for CRAIC
#########################

#remove later (too many errors if done now)
def read_CRAIC(file_path, header_lines=10, print_header=True):
    header = []
    with open(file_path) as file:
        for _ in range(header_lines):
            # file.readline()
            header.append(file.readline().strip())
        wavelengths_line = file.readline().strip().split('\t')
        intensities_line = file.readline().strip().split('\t')
    wavelengths = [float(w) for w in wavelengths_line]
    intensities = [float(i) for i in intensities_line]

    data = pd.DataFrame({'Wavelength': wavelengths, 'Intensity': intensities})

    if print_header is True:
        print('Header Lines:')

        for line in header:
            print(line)
    return data

def CRAIC_map(
    folder,
    background,
    reflection_transmission_list,
    grid,
    unit='nm',
):
    # x axis is taken from the background file, maybe check that it is always the same
    #redo!!!! it relies on folder structure!!!
    reflection_name, transmission_name = reflection_transmission_list
    data = pd.DataFrame()
    npoints = len(grid)
    background = read_CRAIC(os.path.join(folder, background), print_header=False)
    fig, ax = plt.subplots(3, 1, figsize=(10, 10))

    if unit == 'nm':
        x_axis = background['Wavelength']
        x_label = 'Wavelength (nm)'

    elif unit == 'eV':
        h = 4.135 * 10**-15
        c = 3 * 10**8
        wavelength_ev = (h * c) / (background['Wavelength'] * 10**-9)

        x_axis = wavelength_ev
        x_label = 'Energy (eV)'

    for i in range(1, npoints + 1):
        file_refl = f'{reflection_name}-{i}.msp'
        file_transl = f'{transmission_name}-{i}.msp'

        data_R = read_CRAIC(os.path.join(folder, file_refl), print_header=False)
        data_T = read_CRAIC(os.path.join(folder, file_transl), print_header=False)

        data_R['Intensity'] = data_R['Intensity'] - background['Intensity']
        # calculate absorption coefficient
        data_A = -(np.log(data_T['Intensity'] / (100 - data_R['Intensity']))) * 10**5


        ax[0].plot(x_axis, data_R['Intensity'], label=f' {i}')
        ax[1].plot(x_axis, data_T['Intensity'], label=f' {i}')
        ax[2].plot(x_axis, data_A, label=f' {i}')

        ax[0].legend(bbox_to_anchor=(1, 1), loc='upper right')
        ax[1].legend(bbox_to_anchor=(1, 1), loc='upper right')
        ax[2].legend(bbox_to_anchor=(1, 1), loc='upper right')

        ax[0].set_ylabel('R %')
        ax[1].set_ylabel('T %')
        ax[2].set_ylabel(r'$\alpha$ (cm$^{-1}$)')

        for ax_ in ax:
            ax_.set_xlabel(x_label)

        plt.tight_layout()


        # save data in a multiindex dataframe

        data_R['Intensity'].rename('R', inplace=True)
        data_T['Intensity'].rename('T', inplace=True)
        data_A.rename('A', inplace=True)
        data = pd.concat([data, data_R['Wavelength']], axis=1, ignore_index=False)
        data = pd.concat([data, data_R['Intensity']], axis=1, ignore_index=False)
        data = pd.concat([data, data_T['Intensity']], axis=1, ignore_index=False)
        data = pd.concat([data, data_A], axis=1, ignore_index=False)

    coord_header = grid_to_MIheader(
        grid
    )  # check how the points are collected and build the grid accordingly
    df_header = pd.MultiIndex.from_product(
        [coord_header, data.columns[0:4]], names=['Coordinate', 'Data type']
    )
    data_MI = pd.DataFrame(data.values, columns=df_header)

    return data_MI



###########################
# leftover/unused functions
###########################

# def raman_fit(
#     data,
#     Peaks,
#     dataRange,
#     knots,
#     remove_background_fit=False,
# ):
#     """Fit data using models from lmfit. Gaussian for peaks, based on thePeaks output
#     from initial_peaks, and
#     spline background model adjustable with knots. withplots allows
#     toggling plots on/off and picking scale.
#     Outputs: dataframe with Raman shift, measured intensity, fit intensity, peak
#     locations, intensity, and FWHM"""
#     dataRangeMin = dataRange[0]
#     dataRangeMax = dataRange[1]
#     plotscale='log',
#     # setup data
#     column_headers = data.columns.values
#     col_theta = column_headers[::2]
#     col_counts = column_headers[1::2]
#     data = data.iloc[dataRangeMin:dataRangeMax]
#     data.reset_index(drop=True, inplace=True)

#     # empty frame for XRD output
#     RamanoutFrame = pd.DataFrame()

#     # fit all the things number 2
#     for i in range(0, len(col_theta)):
#         # select data
#         x = data[col_theta[i]]
#         y = data[col_counts[i]]

#         # select peaks and remove nans
#         thesePeaks = Peaks[[col_theta[i], col_counts[i]]].dropna()

#         # define peak model
#         mod = None
#         peakNames = []
#         params = Parameters()
#         for idx in range(len(thesePeaks)):
#             this_mod = make_model_raman(
#                 idx, i, thesePeaks, col_counts, col_theta, params)
#             if mod is None:
#                 mod = this_mod
#             else:
#                 mod = mod + this_mod
#             peakNames.append(this_mod.prefix)

#         # define background model
#         knot_xvals = np.linspace(min(x), max(x), knots)
#         bkg = SplineModel(prefix='bkg_', xknots=knot_xvals)
#         params = params.update(bkg.guess(y, x=x))

#         # construct model
#         mod = mod + bkg

#         # fit
#         out = mod.fit(y, params, x=x)
#         comps = out.eval_components(x=x)

#         # extract peak data from fit
#         peakHeights = np.array([])
#         peakCenters = np.array([])
#         peakFWHMs = np.array([])
#         for j in range(len(peakNames)):
#             peakCenter = round(out.params[peakNames[j] + 'center'].value, 2)
#             peakHeight = round(out.params[peakNames[j] + 'height'].value, 3)
#             peakHeights = np.append(peakHeights, peakHeight)
#             peakCenters = np.append(peakCenters, peakCenter)
#             peakFWHMs = np.append(
#                 peakFWHMs, round(out.params[peakNames[j] + 'fwhm'].value, 2)
#             )

#         peakData = np.vstack((peakCenters, peakHeights, peakFWHMs)).T
#         Raman_peaks_header = pd.MultiIndex.from_product(
#             [[col_theta[i][0]], ['Peak Raman shift', 'Peak intensity', 'FWHM']],
#             names=['Coordinate', 'Data type'],
#         )
#         peakOutput = pd.DataFrame(data=peakData, columns=Raman_peaks_header)

#         # extract fit and theta
#         XRD_data_header = pd.MultiIndex.from_product(
#             [[col_theta[i][0]],['Raman shift', 'Measured intensity','Fit intensity']],
#             names=['Coordinate', 'Data type'],
#         )
#         if remove_background_fit is not False:
#             fitData = np.vstack((x, y, out.best_fit - comps['bkg_'])).T
#         else:
#             fitData = np.vstack((x, y, out.best_fit)).T
#         # fitOutput = pd.DataFrame(data=fitData, columns=['2θ','Measured intensity',
#         # 'Fit intensity'])
#         fitOutput = pd.DataFrame(data=fitData, columns=XRD_data_header)
#         RamanoutFrame = pd.concat([RamanoutFrame, fitOutput, peakOutput], axis=1)

#         # plot fit
#         plt.plot(x, y, label='data')
#         plt.plot(x, out.best_fit, label='best fit')
#         plt.xlabel(col_theta[i][1])
#         plt.ylabel(col_counts[i][1])
#         plt.title(col_counts[i][0])
#         plt.yscale(plotscale)
#         plt.legend()
#         plt.show()
#         # print output peak positions, intensity, and FWHM
#         print('Peak positions:\n', peakOutput)

#     return RamanoutFrame


# def old_UPS_fit(data, startvalue, fit_background=True, plotscale='linear'):
#     """Fit UPS data using dataframe input from read_UPS, the data is fitted by finding
#     background and valence bands based on slope, they are both fitted
#     Output: dataframe with the background and fit intersection point, as well as the
#     backgrounds intersect with x"""
#     column_headers = data.columns.values
#     col_BE = column_headers[::2]
#     col_counts = column_headers[1::2]
#     mod = LinearModel(prefix='reg_')

#     UPS_outframe = pd.DataFrame()
#     for i in range(0, len(col_BE)):
#         x = data[col_BE[i]]
#         y = data[col_counts[i]]

#         # reversing the data
#         x_reversed = x[::-1].values
#         y_reversed = y[::-1].values
#         # y_reversed = savgol_filter(y[::-1].values, 2, 1)

#         # select based on input start value
#         xselect = x_reversed[list(range(startvalue, len(x)))]
#         yselect = y_reversed[list(range(startvalue, len(y)))]

#         # finding the background onset from slope
#         k = 1
#         slope_cutof= 2000
#         slope = None
#         while k < len(yselect) - 1:
#             slope = yselect[k] - yselect[k - 1]
#             if slope < slope_cutof:
#                 k = k + 1
#             else:
#                 bkg_start = k
#                 bkg_slope = slope
#                 k = len(yselect) - 1

#         # finding the background end from onset point and slope
#         xselect1 = xselect[list(range(bkg_start, len(xselect)))]
#         yselect1 = yselect[list(range(bkg_start, len(yselect)))]

#         k = 25
#         v=100
#         slope = None
#         while k < v:
#             slope = yselect1[k] - yselect1[k - 1]
#             if slope < bkg_slope * 1.5:
#                 k = k + 1
#             else:
#                 bkg_end = k
#                 k = 100

#         # fitting the background with a linear model
#         x_bkg = xselect1[list(range(0, bkg_end))]
#         y_bkg = yselect1[list(range(0, bkg_end))]
#         out_bkg = mod.fit(y_bkg, x=x_bkg)
#         bkg_x_intercept = (
#             -out_bkg.params['reg_intercept'].value / out_bkg.params['reg_slope'].value
#         )

#         if fit_background is True:
#             # finding the valence onset

#             xselect_valence = x_reversed[
#                 list(range(bkg_end + bkg_start + startvalue, len(x)))
#             ]
#             yselect_valence = y_reversed[
#                 list(range(bkg_end + bkg_start + startvalue, len(y)))
#             ]

#             k = 5
#             slope = None
#             while k < len(yselect_valence) - 1:
#                 slope = yselect_valence[k] - yselect_valence[k - 1]
#                 if slope < bkg_slope * 4:
#                     k = k + 1
#                 else:
#                     valence_start = k
#                     valence_slope = slope
#                     k = len(yselect_valence) - 1

#             # finding valence end
#             xselect_valence1 = xselect_valence[
#                 list(range(valence_start, len(xselect_valence)))
#             ]
#             yselect_valence1 = yselect_valence[
#                 list(range(valence_start, len(yselect_valence)))
#             ]

#             k = 15
#             slope = None
#             while k < len(yselect_valence1) - 1:
#                 slope = yselect_valence1[k] - yselect_valence1[k - 1]
#                 if slope < valence_slope * 1.2:
#                     k = k + 1
#                 else:
#                     valence_end = k
#                     k = len(yselect_valence1) - 1

#             x_valence = xselect_valence1[list(range(0, valence_end))]
#             y_valence = yselect_valence1[list(range(0, valence_end))]
#             out_valence = mod.fit(y_valence, x=x_valence)

#             fits_intercept = (
#                 out_valence.params['reg_intercept'].value
#                 - out_bkg.params['reg_intercept'].value
#             ) / (
#                 out_bkg.params['reg_slope'].value
#                 - out_valence.params['reg_slope'].value
#             )

#             plt.plot(x, y, label='data')
#             plt.plot(x_valence, out_valence.best_fit, label='valence fit')
#             plt.plot(
#                 xselect_valence[valence_start],
#                 yselect_valence[valence_start],
#                 'o',
#                 label='valence_start',
#             )
#             plt.plot(
#                 xselect_valence1[valence_end],
#                 yselect_valence1[valence_end],
#                 'o',
#                 label='valence_end',
#             )
#             plt.plot(xselect[bkg_start], yselect[bkg_start], 'o', label='bkg_start')
#             plt.plot(xselect1[bkg_end], yselect1[bkg_end], 'o', label='bkg_end')
#             plt.plot(x_bkg, out_bkg.best_fit, label='background fit')
#             plt.xlabel(col_BE[i][1])
#             plt.ylabel(col_counts[i][1])
#             plt.title(col_counts[i][0])
#             plt.yscale(plotscale)
#             plt.xlim(-4, 8)
#             plt.ylim(-1e4, 2e6)
#             plt.legend()
#             plt.show()

#             print('bkg_x_intercept:\n', bkg_x_intercept)
#             print('fits_intercept:\n', fits_intercept)

#             intercepts = np.vstack((bkg_x_intercept, fits_intercept)).T
#             UPS_header = pd.MultiIndex.from_product(
#                 [[col_BE[i][0]], ['bkg_x_intercept', 'fits_intercept']],
#                 names=['Coordinate', 'Data type'],
#             )
#             UPS_output = pd.DataFrame(data=intercepts, columns=UPS_header)

#         else:

#             plt.plot(x, y, label='data')
#             plt.plot(x_bkg, out_bkg.best_fit, label='valence fit')
#             plt.xlabel(col_BE[i][1])
#             plt.ylabel(col_counts[i][1])
#             plt.title(col_counts[i][0])
#             plt.yscale(plotscale)
#             plt.xlim(0, 8)
#             plt.ylim(-1e4, 2e6)
#             plt.legend()
#             plt.show()

#             print('bkg_x_intercept:\n', bkg_x_intercept)

#             UPS_header = pd.MultiIndex.from_product(
#                 [[col_BE[i][0]],['bkg_x_intercept']],names=['Coordinate', 'Data type']
#             )
#             UPS_output = pd.DataFrame(
#                 data=np.array([bkg_x_intercept]), columns=UPS_header
#             )

#         UPS_outframe = pd.concat([UPS_outframe, UPS_output], axis=1)
#     return UPS_outframe