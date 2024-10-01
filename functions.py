#%%
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import re
import os
from scipy.interpolate import griddata
from scipy.signal import find_peaks
import seaborn as sns
from lmfit.models import PseudoVoigtModel, SplineModel,LinearModel, GaussianModel
from lmfit import Parameters
from scipy.signal import savgol_filter
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import pickle
from openpyxl import load_workbook


##########################
# Functions related to loading data
##########################

def read_XRD(filename, grid, n = 0, separator = "\t"):
    '''"Read data from an XRD datafile into a dataframe. The data is constrained to a custom grid which must be provided via the "measurement_grid" function."
    Usage: data, coords = read_XRD(filename) Optional: "n" - amount of measurements to include. "separator" - csv file separator.'''
    # read data and limit length based on amount of wanted points
    data = pd.read_csv(filename, sep=separator, header=1)
    if n > 0:
        data = data.iloc[:,0:n*2]
    data.rename(columns={"2θ, °": "2θ (°)", "Intensity, counts": "Intensity (counts)"}, inplace=True)

    # we need coords for aligning data to grid
    # only load row of measurement names, and convert to an array of those names
    file_header = pd.read_csv(filename, sep=separator, header=0, nrows=0)
    coords_array = file_header.columns.values[::2]
    # limit length based on amount of wanted points
    if n > 0:
        coords_array = coords_array[0:n]
    
    # extract coordinate info from headers
    for i in range(len(coords_array)):
        # split header and select coordinates
        split_list = re.split('_', coords_array[i])
        coords_array[i] = split_list[-2:]

        # replace '-' character with '.', but preserve '-' at start for negative numbers
        for j in range(2):
            coords_array[i][j] = re.sub('(?!^)-', '.', coords_array[i][j])

    # convert array to a list otherwise Pandas does not work
    coords_list = list(coords_array)
    coords = pd.DataFrame(coords_list, columns=['x', 'y'])

    # do some treatment on the dataframe
    coords = coords.astype(float)

    # align data to grid
    coordgrid = coords_to_grid(coords, grid)
    coord_header = grid_to_MIheader(coordgrid)

    # construct XRD dataframe with multiindexing for coordinates
    header = pd.MultiIndex.from_product([coord_header, data.columns[0:2]],names=['Coordinate','Data type'])
    data = pd.DataFrame(data.values, columns=header)
    return data, coords

def read_ellipsometry_thickness(filename, grid, n = 0, separator = "\t"):
    '''"Read thickness data and coordinates from an ellipsometry datafile. The data is constrained to a custom grid which must be provided via the "measurement_grid" function."
    Usage: data, coords = read_ellipsometry_thickness(filename) Optional: "n" - amount of measurements to include. "separator" - csv file separator.'''
    # read data and limit length based on amount of wanted points
    data = pd.read_csv(filename, sep=separator, header=1)
    if n > 0:
        data = data.truncate(after=n-1)
    data.rename(columns={"Z": "Z (nm)"}, inplace=True)

    # we need coords for aligning data to grid
    # extract coordinates
    coords = data.copy()
    coords = coords.drop(columns=['Z (nm)'])
    coords.rename(columns={"X (cm)": "x", "Y (cm)": "y"}, inplace=True)
    # convert to float
    coords = coords.astype(float)
    # convert from cm to mm
    coords = coords*10

    # align data to grid
    coordgrid = coords_to_grid(coords, grid)
    coord_header = grid_to_MIheader(coordgrid)

    # construct ellipsometry dataframe with multiindexing for coordinates
    data = data.drop(columns=['X (cm)','Y (cm)'])
    data = data.stack().to_frame().T
    # "verify_integrity = False" lmao
    data.columns = data.columns.set_levels(coord_header, level=0, verify_integrity=False)
    data.columns.rename(["Coordinate", "Data type"], level=[0, 1], inplace = True)
    return data, coords

def read_ellipsometry_MSE(filename, grid, n = 0, separator = "\t"):
    '''"Read Mean Squared Error data and coordinates from an ellipsometry datafile. The data is constrained to a custom grid which must be provided via the "measurement_grid" function."
    Usage: data, coords = read_ellipsometry_MSE(filename) Optional: "n" - amount of measurements to include. "separator" - csv file separator.'''
    # read data and limit length based on amount of wanted points
    data = pd.read_csv(filename, sep=separator, header=1)
    if n > 0:
        data = data.truncate(after=n-1)
    data.rename(columns={"Z": "MSE"}, inplace=True)

    # we need coords for aligning data to grid
    # extract coordinates
    coords = data.copy()
    coords = coords.drop(columns=['MSE'])
    coords.rename(columns={"X (cm)": "x", "Y (cm)": "y"}, inplace=True)
    # convert to float
    coords = coords.astype(float)
    # convert from cm to mm
    coords = coords*10

    # align data to grid
    coordgrid = coords_to_grid(coords, grid)
    coord_header = grid_to_MIheader(coordgrid)

    # construct ellipsometry dataframe with multiindexing for coordinates
    data = data.drop(columns=['X (cm)','Y (cm)'])
    data = data.stack().to_frame().T
    # "verify_integrity = False" lmao
    data.columns = data.columns.set_levels(coord_header, level=0, verify_integrity=False)
    data.columns.rename(["Coordinate", "Data type"], level=[0, 1], inplace = True)
    return data, coords

def read_ellipsometry_nk(filename, grid, n = 0, separator = "\t"):
    '''"Read refractive index n and absorption coefficient k data and coordinates from an ellipsometry datafile. The data is constrained to a custom grid which must be provided via the "measurement_grid" function."
    Usage: data, coords = read_ellipsometry_nk(filename) Optional: "n" - amount of measurements to include. "separator" - csv file separator.'''
    # read data and split into energy and n/k data, limited by number of wanted points
    data = pd.read_csv(filename, sep=separator, header=1, index_col = False)
    data_energy = data.iloc[:,0]
    if n > 0:
        data_n_k = data.iloc[:,1:n*2+1]
    else:
        data_n_k = data.iloc[:,1:]

    # get headers from n/k data to get an array of coordinates
    coords_array = np.array(data_n_k.columns)

    # extract coordinate info from headers
    for i in range(len(coords_array)):
        # split header and select coordinates
        split_list = re.split(',', coords_array[i])
        split_list[0] = float(split_list[0][4:])
        split_list[1] = float(split_list[1][:-1])
        coords_array[i] = split_list
    coords_array = coords_array[::2]

    # convert array to a list otherwise Pandas does not work
    coords_list = list(coords_array)
    coords = pd.DataFrame(coords_list, columns=['x', 'y'])

    # convert from cm to mm
    coords = coords*10

    # align data to grid
    coordgrid = coords_to_grid(coords, grid)
    coord_header = grid_to_MIheader(coordgrid)

    # rename all n and k columns, and insert energy column before each n and k set
    k = 0
    for i in range(0, len(data_n_k.columns), 2):
        data_n_k.columns.values[i+k] = "n"
        data_n_k.columns.values[i+k+1] = "k"
        data_n_k.insert(loc=i+k, column='Energy (eV)'.format(i), value=data_energy, allow_duplicates = True)
        k += 1

    # construct XRD dataframe with multiindexing for coordinates
    header = pd.MultiIndex.from_product([coord_header, data.columns[0:3]],names=['Coordinate','Data type'])
    data = pd.DataFrame(data_n_k.values, columns=header)
    return data, coords

def convert_to_eV(data):
    '''"Convert ellipsometry data in wavelength to eV."'''
    # Planck's constant (eV/Hz)
    h = 4.135*10**-15
    # Speed of light (m/s)
    c = 3*10**8
    data = data.copy()
    data.iloc[:, data.columns.get_level_values(1) == 'Wavelength (nm)'] = (h*c)/(data.iloc[:, data.columns.get_level_values(1) == 'Wavelength (nm)']*10**-9)
    data.columns = data.columns.set_levels(['Energy (eV)', 'k', 'n'], level=1)
    data = data.round(3)
    return data

def read_layerprobe_coords(filename, n = 0, separator = "\t"):
    '''"Read only the coordinates of a set of LayerProbe measurements. Necessary for interpolation until I decide to fix it :)"'''
    # read data and limit length based on amount of wanted points
    data = pd.read_csv(filename, sep=separator, header=0)
    if n > 0:
        data = data.truncate(after=n-1)
    data = data.sort_values(by=['X (mm)','Y (mm)'])
    data = data.reset_index(drop = True)
    # extract coordinates
    coords = data.iloc[:,1:3]
    coords.rename(columns={"X (mm)": "x", "Y (mm)": "y"}, inplace=True)
    coords = coords.astype(float)
    return coords


def read_layerprobe(filename, grid, sheetname = -1, n = 0):
    '''"Read data and coordinates from a LayerProbe datafile. The file should be an Excel sheet. The data is constrained to a custom grid which must be provided via the "measurement_grid" function."
    Usage: data, coords = read_layerprobe(filename) Optional: "sheetname" name of sheet, defaults to last sheet in file. "n" - amount of measurements to include.'''
   # read data and limit length based on amount of wanted points
    data = pd.read_excel(filename, sheet_name = sheetname, header=0)
    if n > 0:
        data = data.truncate(after=n-1)
    data = data.sort_values(by=['X (mm)','Y (mm)'])
    data = data.reset_index(drop = True)
    # we need coords for aligning data to grid
    # extract coordinates
    coords = data.copy()
    coords = coords.iloc[:,1:3]
    coords.rename(columns={"X (mm)": "x", "Y (mm)": "y"}, inplace=True)
    # treat coordinates
    coords = coords.astype(float)
    coords = coords.round(4)

    # remove coordinates from data
    data = data.drop(data.columns[0:3], axis = 1)

    # align data to grid
    coordgrid = coords_to_grid(coords, grid)
    coord_header = grid_to_MIheader(coordgrid)

    # construct dataframe with multiindexing for coordinates
    header = pd.MultiIndex.from_product([coord_header, data],names=['Coordinate','Data type'])
    data = pd.DataFrame(data.to_numpy().flatten()).transpose()
    data = pd.DataFrame(data.values, columns=header)
    return data, coords

def read_XPS(filename, grid):
    '''"Read data and coordinates from an XPS datafile. The file should be an csv (.txt) file. The data is constrained to a custom grid which must be provided via the "measurement_grid" function."
    Usage: data, coords = read_XPS(filename, grid)"'''
    # read the file
    file = pd.read_csv(filename, encoding = 'ANSI', engine='python', sep='delimiter', header = None, skiprows = 29)
    file.drop(file.iloc[4::7].index, inplace=True)
    file.reset_index(drop = True)

    # the file has a really weird format so we need to do a lot of work to extract data
    # get amount of peaks
    peaknumb = []
    for i in range(0, len(file), 6):
        peaknumb.append(int(file.iloc[i][0].split()[8].replace(";","")))
    n = max(peaknumb) + 1

    # remove useless rows
    file.drop(file.iloc[0::6].index, inplace=True)
    file.reset_index(drop = True)

    # get data from remaining rows
    full_peaklist = []
    peaklist = []
    coordlist = []
    datalist = []
    for i in range(0, len(file), 5):
        # load peak type and coordinates and fix formatting
        peaktype = ' '.join(file.iloc[i][0].split()[5:len(file.iloc[i][0].split())]).replace("VALUE='","").replace("';","")
        xcoord = float(file.iloc[i+1][0].split()[5].replace("VALUE=","").replace(";",""))
        ycoord = float(file.iloc[i+2][0].split()[5].replace("VALUE=","").replace(";",""))
        coords = [xcoord, ycoord]
        # load data
        data = file.iloc[i+3][0].split()[2::]
        data.append(file.iloc[i+4][0].split()[2::][0])
        # fix data formatting
        data = [j.replace(",","") for j in data]
        data = [round(float(j),3) for j in data]

        full_peaklist.append(peaktype)
        peaklist.append(peaktype.split()[0])
        coordlist.append(coords)
        datalist.append(data)

    # create data dataframe
    dataframe = pd.DataFrame(datalist, columns = ['Intensity (counts)','Atomic %','Area (counts*eV)','FWHM (eV)','Peak BE (eV)'])
    # modify some values
    # convert cps to counts (machine does 25 cps)
    dataframe['Intensity (counts)'] = dataframe['Intensity (counts)']/25
    # convert KE to BE (KE of machine X-rays is 1486.68 eV)
    dataframe['Peak BE (eV)'] = 1486.68 - dataframe['Peak BE (eV)']
    # reorder columns to be similar to Avantage
    columnorder = ['Peak BE (eV)','Intensity (counts)','FWHM (eV)','Area (counts*eV)','Atomic %']
    dataframe = dataframe.reindex(columnorder, axis=1)

    # create coordinate dataframe
    coords = pd.DataFrame(coordlist, columns=['x', 'y'])
    # remove duplicate coordinates
    coords = coords.drop_duplicates(ignore_index = True)
    # adjust range to center coords on 0,0 instead of upper left corner
    coords['x'] = coords['x'] - max(coords['x'])/2
    coords['y'] = coords['y'] - max(coords['y'])/2
    # convert coords from µm to mm
    coords = coords/1000
    # flip y coordinate because Avantage is mental
    coords['y'] = coords['y'].values[::-1]

    # create peak dataframe
    peaks = pd.DataFrame(peaklist, columns = ['Peak'])
    # add peak dataframe to front of data dataframe
    dataframe = pd.concat([peaks, dataframe], axis = 1)

    # add column with summed atomic %
    element_list = dataframe['Peak'].unique()
    atomic_percent_list = []
    for l in range(0, int(len(peaklist)), n):
        for k in range(len(element_list)):
            atomic_percent = round(sum(dataframe.iloc[l:l+n].loc[dataframe['Peak'] == element_list[k]]["Atomic %"]),3)
            atomic_percent_list.append(atomic_percent)
        for j in range(len(element_list)*(n-1)):
            atomic_percent_list.append(float("NaN"))
    atomic_percent_array = np.split(np.array(atomic_percent_list), len(atomic_percent_list)/len(element_list))
    atomic_percent_frame = pd.DataFrame(atomic_percent_array, columns=element_list + " Total")
    dataframe = pd.concat([dataframe, atomic_percent_frame], axis=1)

    # align data to grid
    coordgrid = coords_to_grid(coords, grid)
    coord_header = grid_to_MIheader(coordgrid)

    # construct XRD dataframe with multiindexing for coordinates
    header = pd.MultiIndex.from_product([coord_header, dataframe.columns],names=['Coordinate','Data type'])
    # reorder dataframe stacking to fit coordinate attachment
    n2 = n
    stackedframe = np.hstack([dataframe.values[0:n2],(dataframe.values[n2:2*n2])])
    for i in range(2*n2, len(dataframe), n2):
        stackedframe = np.hstack([stackedframe, (dataframe.values[i:i+n2])])
    data = pd.DataFrame(stackedframe, columns=header)
    return data, coords

def read_UPS(filename, grid):
    '''"Read data and coordinates from an UPS datafile. The file should be an Excel (.xlsx) file. The data is constrained to a custom grid which must be provided via the "measurement_grid" function."
    Usage: data, coords = read_UPS(filename, grid)"'''
    # load data, energy, and coordinates from all sheets
    dataload_counts = pd.read_excel(filename, sheet_name = None, skiprows = 18)
    dataload_eV = pd.read_excel(filename, sheet_name = None, skiprows = 18, usecols = [0])
    coordload = pd.read_excel(filename, sheet_name = None, skiprows = 13, nrows = 3)

    # select dictionary keys for only usable sheets
    dictlist = list(dataload_eV.keys())

    # remove useless "Titles" sheets that Avantage generates
    # also remove "Peak Table" sheets as we get coordinates from graph sheets
    j = 0
    for i in range(len(dictlist)):
        if dictlist[j].startswith("Titles"):
            del dictlist[j]
            j -= 1
        if dictlist[j].startswith("Peak Table"):
            del dictlist[j]
            j -= 1
        j += 1

    # remove last sheet in file as it should be a blank sheet
    del dictlist[-1]

    data_list = []
    xy_coords_list = []

    # read data
    for i in range(0, len(dictlist)):
        # load data from usable sheets only
        dataselect_counts = dataload_counts[dictlist[i]].dropna(axis = 1, how = 'all')
        data = dataselect_counts.iloc[:,1:]
        data_eV = dataload_eV[dictlist[i]]
        # rename columns
        data.columns.values[:] = "Intensity (counts)"
        # insert energy column besides each data column
        j = 0
        for k in range(0, len(data.columns)):
            data.insert(loc=k+j, column='BE (eV)'.format(k), value=data_eV, allow_duplicates = True)
            j += 1
        data_list.append(data)

        # read coords
        coordselect = coordload[dictlist[i]].dropna(axis=1, how = 'all')
        xcoord = coordselect.iloc[0,1]
        ycoord = coordselect.iloc[2,2:]
        # create coords list
        for l in range(len(ycoord)):
            x = xcoord
            y = ycoord[l]
            xy_coords = np.array([x, y])
            xy_coords_list.append(xy_coords)

    # create coords dataframe
    coords = pd.DataFrame(xy_coords_list, columns=['x', 'y'])

    # create merged data dataframe
    dataframe = pd.concat(data_list, axis = 1)

    # adjust range to center coords on 0,0 instead of upper left corner
    coords['x'] = coords['x'] - max(coords['x'])/2
    coords['y'] = coords['y'] - max(coords['y'])/2

    # convert coords from µm to mm
    coords = coords/1000

    # align data to grid
    coordgrid = coords_to_grid(coords, grid)
    coord_header = grid_to_MIheader(coordgrid)

    # construct XRD dataframe with multiindexing for coordinates
    header = pd.MultiIndex.from_product([coord_header, data.columns.unique()],names=['Coordinate','Data type'])
    data = pd.DataFrame(dataframe.values, columns=header)
    return data, coords

def read_REELS(filename, grid):
    '''"Read data and coordinates from an REELS datafile. The file should be an Excel (.xlsx) file. The data is constrained to a custom grid which must be provided via the "measurement_grid" function."
    Usage: data, coords = read_REELS(filename, grid)"'''
    # This data loading method is identical to UPS.
    data, coords = read_UPS(filename, grid)
    return data, coords

def read_raman(filename, grid):
    '''"Read data and coordinates from a Raman spectroscopy datafile. The file should be an Excel (.xlsx) file. The data is constrained to a custom grid which must be provided via the "measurement_grid" function."
    Usage: data, coords = read_raman(filename)"'''
    # load data, energy, and coordinates from all sheets
    dataload_counts = pd.read_excel(filename, sheet_name = None, skiprows = 18)
    dataload_eV = pd.read_excel(filename, sheet_name = None, skiprows = 18, usecols = [0])
    coordload = pd.read_excel(filename, sheet_name = None, skiprows = 13, nrows = 3)

    # select dictionary keys for only usable sheets
    dictlist = list(dataload_eV.keys())

    # remove useless "Titles" sheets that Avantage generates
    # also remove "Peak Table" sheets as we get coordinates from graph sheets
    j = 0
    for i in range(len(dictlist)):
        if dictlist[j].startswith("Titles"):
            del dictlist[j]
            j -= 1
        if dictlist[j].startswith("Peak Table"):
            del dictlist[j]
            j -= 1
        j += 1

    # remove last sheet in file as it should be a blank sheet
    del dictlist[-1]

    data_list = []
    xy_coords_list = []

    # read data
    for i in range(0, len(dictlist)):
        # load data from usable sheets only
        data = dataload_counts[dictlist[i]].iloc[:,2:]
        data_eV = dataload_eV[dictlist[i]]
        # rename columns
        data.columns.values[:] = "Intensity (counts)"
        # insert energy column besides each data column
        j = 0
        for k in range(0, len(data.columns)):
            data.insert(loc=k+j, column='Raman shift (cm^-1)'.format(k), value=data_eV, allow_duplicates = True)
            j += 1
        data_list.append(data)

        # read coords
        xcoord = coordload[dictlist[i]].iloc[0,1]
        ycoord = coordload[dictlist[i]].iloc[2,2:]
        # create coords list
        for l in range(len(ycoord)):
            x = xcoord
            y = ycoord[l]
            xy_coords = np.array([x, y])
            xy_coords_list.append(xy_coords)

    # create coords dataframe
    coords = pd.DataFrame(xy_coords_list, columns=['x', 'y'])

    # create merged data dataframe
    dataframe = pd.concat(data_list, axis = 1)

    # adjust range to center coords on 0,0 instead of upper left corner
    coords['x'] = coords['x'] - max(coords['x'])/2
    coords['y'] = coords['y'] - max(coords['y'])/2

    # convert coords from µm to mm
    coords = coords/1000

    # align data to grid
    coordgrid = coords_to_grid(coords, grid)
    coord_header = grid_to_MIheader(coordgrid)

    # construct XRD dataframe with multiindexing for coordinates
    header = pd.MultiIndex.from_product([coord_header, data.columns.unique()],names=['Coordinate','Data type'])
    data = pd.DataFrame(dataframe.values, columns=header)
    return data, coords

##########################
# Functions related to constructing a grid and inserting data into it
##########################

def measurement_grid(ncolumns, nrows, gridlength, gridheight, startcolumn = 0, startrow = 0):
    '''"Define a grid based on number of columns and rows, length and height of grid in mm, and the first coordinate (lower left corner) in the column and row."'''
    xgrid = np.round(np.linspace(startcolumn, gridlength+startcolumn, ncolumns), 3)
    ygrid = np.round(np.linspace(startrow, gridheight+startrow, nrows), 3)
    grid = np.array([xgrid[0], ygrid[0]])
    for i in range(len(xgrid)):
        for j in range(len(ygrid)):
            grid = np.vstack((grid, np.array([xgrid[i], ygrid[j]])))
    grid = grid[1:]
    grid = pd.DataFrame(grid, columns = ['x','y'])
    return grid

def coords_to_grid(coords, grid):
    '''"Constrain a set of measured datapoints to a custom defined grid made with the "measurement_grid" function."'''
    griddata = coords.copy()
    for i in range(len(coords)):
        # find closest x and y coordinate
        xminindex = np.abs(grid - coords.iloc[i,:]).idxmin().iloc[0]
        yminindex = np.abs(grid - coords.iloc[i,:]).idxmin().iloc[1]
        # assign new coordinates
        griddata.iloc[i,0] = np.round(grid.iloc[xminindex, 0], 2)
        griddata.iloc[i,1] = np.round(grid.iloc[yminindex, 1], 2)
    return(griddata)

def grid_to_MIheader(grid):
    '''"Convert a grid (array of x,y) into a multi index header"'''
    MIgrid = []
    for i in range(len(grid)):
        MIgrid = np.append(MIgrid, ('{},{}'.format(grid.iloc[i,0], grid.iloc[i,1])))
    return MIgrid

def MI_to_grid(MIgrid):
    '''"Convert multi index into a grid (array of x,y)"'''
    MIgrid = MIgrid.columns.get_level_values(0)
    splitvals = re.split(',', MIgrid[0])
    grid = np.array([splitvals[0], splitvals[1]])
    for i in range(1, len(MIgrid)):
        splitvals = re.split(',', MIgrid[i])
        grid = np.vstack((grid, np.array([splitvals[0], splitvals[1]])))
    grid = pd.DataFrame(grid, columns = ['x','y'])
    grid = grid.astype(float)
    return grid

def interpolate_grid(data, grid):
    '''"Interpolate data over a custom grid made with the "measurement_grid" function."'''
    # !!!
    # specifically remove the "Peak" column, which will be present if loaded data is XPS
    # !!!
    data = data.drop(columns = 'Peak', level=1, errors = 'ignore')

    # get grid-aligned coordinates for datapoints for interpolation
    coords = MI_to_grid(data).drop_duplicates(ignore_index=True)

    # interpolation
    # we have to account for multiple variables for every coordinate
    # this creates a list of the data
    dataT = data.transpose()
    dataN = int(len(dataT)/len(coords))
    interpolated_data = []
    for i in range(dataN):
        interpolated_data.append(griddata(coords, dataT.iloc[i::dataN], grid, method = "cubic"))

    # convert the list of data to an array with columns alternating between data types
    if dataN == 0:
        interpolated_array = np.transpose(interpolated_data)
    else:
        interpolated_array = [None]*len(interpolated_data[0])*dataN
        for i in range(dataN):
            interpolated_array[i::dataN] = interpolated_data[i]
        interpolated_array = np.transpose(interpolated_array)

    # list of column names
    columnlist = []
    for i in range(interpolated_array.shape[1]):
        columnlist.append(data.columns[i%dataN][1])

    # dataframe
    interpolated_frame = pd.DataFrame(interpolated_array, columns = columnlist)

    # convert grid to multiindex header
    coord_header = grid_to_MIheader(grid)

    # construct dataframe with multiindexing for coordinates
    header = pd.MultiIndex.from_product([coord_header, list(interpolated_frame.columns.unique())], names=['Coordinate', 'Data type'])
    interpolated_frame = pd.DataFrame(interpolated_frame.values, columns=header)
    return interpolated_frame


##########################
# Functions related to XRD and Raman analysis
##########################

#peaks model function for XRD
def make_model_xrd(num,i,peaks,col_counts,col_theta,params):
    '''constructs a model for every peak based on peaks output from initial_peaks'''
    pref = "f{0}_".format(num)
    model = PseudoVoigtModel(prefix=pref)
    ypeak = peaks[col_counts[i]][peaks.index[num]]
    xpeak = peaks[col_theta[i]][num].astype(float)
    #width = widths_initial[num]
    params.update(model.make_params(center = dict(value=xpeak, min=xpeak*0.9, max=xpeak*1.1),
                                    amplitude = dict(value=ypeak, min=0.2 * ypeak, max=1.2*ypeak)
                                    ))
    return model
#peaks model function for Raman
def make_model_raman(num,i,peaks,col_counts,col_theta,params):
    '''constructs a model for every peak based on peaks output from initial_peaks'''
    pref = "f{0}_".format(num)
    model = GaussianModel(prefix=pref)
    ypeak = peaks[col_counts[i]][peaks.index[num]]
    xpeak = peaks[col_theta[i]][num].astype(float)
    #width = widths_initial[num]
    params.update(model.make_params(center = dict(value=xpeak, min=xpeak*0.9, max=xpeak*1.1),
                                    height = dict(value=ypeak, min=0.2 * ypeak, max=1.2*ypeak),
                                    ))
    return model

def initial_peaks(data,dataRangeMin, dataRangeMax,filterstrength, peakprominence,peakwidth, withplots = True, plotscale = 'log'):
    '''finds peaks using scipy find_peaks on filtered data to construct a model for fitting, filter strength is based on filtervalue and 
    peak find sensitivity based on peakprominence, withplots and plotscale allows toggling plots on/off and picking scale.
    Output: dataframe with peak locations and intensity, to be used for raman_fit or xrd_fit, and data limited by the dataRangemin/max, in index'''
    #setup data
    column_headers = data.columns.values
    col_theta = column_headers[::2]
    col_counts = column_headers[1::2]
    data = data.iloc[dataRangeMin:dataRangeMax]
    data.reset_index(drop=True, inplace=True)

    #create list for intital peaks
    thePeakss = []
    dataCorrect1 = []

    #finding the peaks
    for i in range(0,len(col_theta)):
        #select data
        dataSelect = data[col_counts[i]].copy()
        x = data[col_theta[i]]
        

        #Filter to avoid fake peaks
        if filterstrength > 0:
            l = filterstrength
            #dataSelect = lfilter(b, a, dataSelect)
            dataSelect = savgol_filter(dataSelect, filterstrength, 1)
        

        #find peaks
        peaks, _ = find_peaks(dataSelect,prominence= peakprominence,width = peakwidth)

        #plot
        if withplots == 1:
            plt.plot(x,dataSelect)
            plt.plot(x[peaks], dataSelect[peaks], 'x')
            plt.yscale(plotscale)
            plt.xlabel(col_theta[i][1])
            plt.ylabel(col_counts[i][1])
            plt.title(col_counts[i][0])
            plt.show()


        #save peaks data in the list
        peaksOut = data[[col_theta[i], col_counts[i]]].loc[peaks]
        peaksOut.reset_index(drop=True, inplace=True)
        thePeakss.append(peaksOut)

        #save peaks data in the list
        dataCorr = np.vstack((x,dataSelect)).T
        #dataCorr = pd.DataFrame(data=dataCorrect, columns=column_headers)
        dataCorrect1.append(dataCorr)

    #convert list to dataframe
    thePeaks = pd.concat(thePeakss, axis=1)
    dataCorrected = np.concatenate(dataCorrect1, axis=1)
    dataCorrected = pd.DataFrame(dataCorrected, columns = data.columns)
    return thePeaks, dataCorrected

def xrd_fit(data,Peaks,dataRangeMin, dataRangeMax,knots, withplots = True,plotscale = 'log',remove_background_fit = False):
    '''Fit data using models from lmfit. Pseudo-voigt for peaks, based on thePeaks output from initial_peaks, and 
    spline background model adjustable with knots. withplots and plotscale allows toggling plots on/off and picking scale. 
    Outputs: dataframe with theta, measured intensity, the entire fit, peak locations, intensity, FWHM, and Lorentzian/Gaussian fraction '''
    #setup data
    column_headers = data.columns.values
    col_theta = column_headers[::2]
    col_counts = column_headers[1::2]
    data = data.iloc[dataRangeMin:dataRangeMax]
    data.reset_index(drop=True, inplace=True)
    
    #empty frame for XRD output
    XRDoutFrame = pd.DataFrame()

    #fit all the things number 2
    for i in range(0,len(col_theta)):
        #select data
        x = data[col_theta[i]]
        y = data[col_counts[i]]

        #select peaks and remove nans
        thesePeaks = Peaks[[col_theta[i],col_counts[i]]].dropna()

        #define peak model
        mod = None
        peakNames = []
        params = Parameters()
        for l in range(len(thesePeaks)):
            this_mod = make_model_xrd(l,i,thesePeaks,col_counts,col_theta,params)
            if mod is None:
                mod = this_mod
            else:
                mod = mod + this_mod
            peakNames.append(this_mod.prefix)


        #define background model
        knot_xvals = np.linspace(min(x), max(x), knots)
        bkg = SplineModel(prefix='bkg_', xknots = knot_xvals)
        params = params.update(bkg.guess(y,x=x))

        #construct model
        mod = mod + bkg

        #fit
        out = mod.fit(y, params, x=x)
        comps = out.eval_components(x=x)    

        #extract peak data from fit
        peakHeights = np.array([])
        peakCenters = np.array([])
        peakFWHMs = np.array([])
        peakFractions = np.array([])
        for j in range(len(peakNames)):
            peakCenter = round(out.params[peakNames[j] + 'center'].value,2)
            peakHeight  = round(out.params[peakNames[j] + 'height'].value,3)
            peakHeights = np.append(peakHeights,peakHeight)
            peakCenters = np.append(peakCenters,peakCenter)
            peakFWHMs = np.append(peakFWHMs,round(out.params[peakNames[j] + 'fwhm'].value,2))
            peakFractions = np.append(peakFractions,round(out.params[peakNames[j] + 'fraction'].value,2))

        peakData = np.vstack((peakCenters,peakHeights,peakFWHMs,peakFractions)).T
        XRD_peaks_header = pd.MultiIndex.from_product([[col_theta[i][0]],['Peak 2θ','Peak intensity','FWHM','Lorentzian/Gaussian fraction']],names=['Coordinate','Data type'])
        columns = ['Peak 2θ','Peak intensity','FWHM','Lorentzian/Gaussian fraction']
        peakOutput = pd.DataFrame(data=peakData, columns=XRD_peaks_header)

        #extract fit and theta
        XRD_data_header = pd.MultiIndex.from_product([[col_theta[i][0]],['2θ','Measured intensity','Fit intensity', 'Background']],names=['Coordinate','Data type'])
        if remove_background_fit != False:
            fitData = np.vstack((x,y,out.best_fit-comps['bkg_'], comps['bkg_'])).T
        else:
            fitData = np.vstack((x,y,out.best_fit, comps['bkg_'])).T
        fitOutput = pd.DataFrame(data=fitData, columns=XRD_data_header)
        XRDoutFrame = pd.concat([XRDoutFrame, fitOutput, peakOutput], axis = 1)

        #plot fit
        if withplots == 1:
            plt.plot(x,y, label = 'data')
            plt.plot(x, out.best_fit, label='best fit')
            plt.plot(x, comps['bkg_'],'--', label='background')
            plt.xlabel(col_theta[i][1])
            plt.ylabel(col_counts[i][1])
            plt.title(col_counts[i][0])
            plt.yscale(plotscale)
            plt.legend()
            plt.show()
            #print output peak positions, intensity, and FWHM
            print("Peak positions:\n",peakOutput)       


    return XRDoutFrame


def XRD_background(data,peaks, cut_range=2, order=4,window_length=10, Si_cut=True, withplots= True ):
    data_out = data.copy()
    headerlength = len(data.columns.get_level_values(1).unique())
    col_theta = data.columns.values[::2]
    col_counts = data.columns.values[1::2]
    peaks_theta = peaks.columns.values[::2]

    k=0

    for i in range(0, len(col_theta)):
        cut_intensity=[]

        two_theta = data[col_theta[i]]
        intensity = data[col_counts[i]]
        idx_range = np.where(two_theta >= 20+cut_range)[0][0]

        # Cut data around peaks
        for j in range(len(intensity)):
            if data[col_theta[i]][j] in peaks[peaks_theta[i]].values:
                start_index = max(0, j-idx_range)
                end_index = min(len(data), j+idx_range)
                data_out[col_counts[i]][start_index:end_index] = np.nan #cut data intensity around peaks in data_out

        if Si_cut==True:
            idx_Si = np.where((two_theta >= 60) & (two_theta<= 70))[0]
            data_out[col_counts[i]][idx_Si] = np.nan

        cut_intensity = data_out[col_counts[i]]

        # Smooth the data for better peak detection
        smoothed_intensity = savgol_filter(intensity, window_length=window_length, polyorder=3)
        # Filter out NaN values (they exist because we cut the data) before fitting
        mask = ~np.isnan(cut_intensity)
        filtered_two_theta = two_theta[mask]
        filtered_intensity = intensity[mask]

        # Perform polynomial fitting with filtered data
        background_poly_coeffs = np.polyfit(filtered_two_theta, filtered_intensity, order)
        background = np.polyval(background_poly_coeffs, two_theta)

        # Subtract background
        corrected_intensity = smoothed_intensity - background

        data_out.insert(headerlength*(i+1)+k, "{}".format(data.columns.get_level_values(0).unique()[i]), background, allow_duplicates=True)
        data_out.rename(columns={'': 'Background'}, inplace = True)
        data_out.insert(headerlength*(i+1)+k+1, "{}".format(data.columns.get_level_values(0).unique()[i]), corrected_intensity, allow_duplicates=True)
        data_out.rename(columns={'': 'Corrected Intensity'}, inplace = True)
        k=k+2

        if withplots==True:
            plt.figure()
            coord= data_out.columns.get_level_values(0).unique()[i]
            plt.plot(two_theta, intensity, label='Original Data')
            plt.plot(filtered_two_theta, filtered_intensity, label='filtered Data')
            plt.plot(two_theta, background, label='Background, order='+str(order), linestyle='--')
            plt.plot(two_theta, corrected_intensity, label='Corrected Data')
            plt.title('XRD data at {}'.format(coord))
            plt.legend()
            plt.show()
    display(data_out)
    
    return data_out

def raman_fit(data,Peaks,dataRangeMin, dataRangeMax,knots, withplots = True,plotscale = 'log',remove_background_fit = False):
    '''Fit data using models from lmfit. Gaussian for peaks, based on thePeaks output from initial_peaks, and 
    spline background model adjustable with knots. withplots and plotscale allows toggling plots on/off and picking scale. 
    Outputs: dataframe with Raman shift, measured intensity, fit intensity, peak locations, intensity, and FWHM '''
    #setup data
    column_headers = data.columns.values
    col_theta = column_headers[::2]
    col_counts = column_headers[1::2]
    data = data.iloc[dataRangeMin:dataRangeMax]
    data.reset_index(drop=True, inplace=True)
    
    #empty frame for XRD output
    RamanoutFrame = pd.DataFrame()

    #fit all the things number 2
    for i in range(0,len(col_theta)):
        #select data
        x = data[col_theta[i]]
        y = data[col_counts[i]]

        #select peaks and remove nans
        thesePeaks = Peaks[[col_theta[i],col_counts[i]]].dropna()

        #define peak model
        mod = None
        peakNames = []
        params = Parameters()
        for l in range(len(thesePeaks)):
            this_mod = make_model_raman(l,i,thesePeaks,col_counts,col_theta,params)
            if mod is None:
                mod = this_mod
            else:
                mod = mod + this_mod
            peakNames.append(this_mod.prefix)


        #define background model
        knot_xvals = np.linspace(min(x), max(x), knots)
        bkg = SplineModel(prefix='bkg_', xknots = knot_xvals)
        params = params.update(bkg.guess(y,x=x))

        #construct model
        mod = mod + bkg

        #fit
        out = mod.fit(y, params, x=x)
        comps = out.eval_components(x=x)    

        #extract peak data from fit
        peakHeights = np.array([])
        peakCenters = np.array([])
        peakFWHMs = np.array([])
        peakFractions = np.array([])
        for j in range(len(peakNames)):
            peakCenter = round(out.params[peakNames[j] + 'center'].value,2)
            peakHeight  = round(out.params[peakNames[j] + 'height'].value,3)
            peakHeights = np.append(peakHeights,peakHeight)
            peakCenters = np.append(peakCenters,peakCenter)
            peakFWHMs = np.append(peakFWHMs,round(out.params[peakNames[j] + 'fwhm'].value,2))

        peakData = np.vstack((peakCenters,peakHeights,peakFWHMs)).T
        Raman_peaks_header = pd.MultiIndex.from_product([[col_theta[i][0]],['Peak Raman shift','Peak intensity','FWHM']],names=['Coordinate','Data type'])
        peakOutput = pd.DataFrame(data=peakData, columns=Raman_peaks_header)

        #extract fit and theta
        XRD_data_header = pd.MultiIndex.from_product([[col_theta[i][0]],['Raman shift','Measured intensity','Fit intensity']],names=['Coordinate','Data type'])
        if remove_background_fit != False:
            fitData = np.vstack((x,y,out.best_fit-comps['bkg_'])).T
        else:
            fitData = np.vstack((x,y,out.best_fit)).T
        #fitOutput = pd.DataFrame(data=fitData, columns=['2θ','Measured intensity','Fit intensity'])
        fitOutput = pd.DataFrame(data=fitData, columns=XRD_data_header)
        RamanoutFrame = pd.concat([RamanoutFrame, fitOutput, peakOutput], axis = 1)

        #plot fit
        if withplots == 1:
            plt.plot(x,y, label = 'data')
            plt.plot(x, out.best_fit, label='best fit')
            plt.xlabel(col_theta[i][1])
            plt.ylabel(col_counts[i][1])
            plt.title(col_counts[i][0])
            plt.yscale(plotscale)
            plt.legend()
            plt.show()
            #print output peak positions, intensity, and FWHM
            print("Peak positions:\n",peakOutput)       


    return RamanoutFrame

##########################
# Functions related to UPS analysis
##########################
def UPS_fit(data,startvalue, fit_background = True, withplots = True,plotscale = 'linear'):
    '''Fit UPS data using dataframe input from read_UPS, the data is fitted by finding background and valence bands based on slope, they are both fitted
    Output: dataframe with the background and fit intersection point, as well as the backgrounds intersect with x'''
    column_headers = data.columns.values
    col_BE = column_headers[::2]
    col_counts = column_headers[1::2]
    mod = LinearModel(prefix='reg_')


    UPS_outframe = pd.DataFrame()
    for i in range(0,len(col_BE)):
        x = data[col_BE[i]]
        y = data[col_counts[i]]

        #reversing the data
        x_reversed = x[::-1].values
        y_reversed = y[::-1].values
        #y_reversed = savgol_filter(y[::-1].values, 2, 1)

        #select based on input start value
        xselect = x_reversed[list(range(startvalue,len(x)))]
        yselect = y_reversed[list(range(startvalue,len(y)))]

        #finding the background onset from slope
        k = 1
        slope = None
        while k < len(yselect)-1:
            slope = yselect[k]-yselect[k-1]
            if slope < 2000:
                k = k + 1
            else:
                bkg_start = k
                bkg_slope = slope
                k = len(yselect)-1

        #finding the background end from onset point and slope
        xselect1 = xselect[list(range(bkg_start,len(xselect)))]
        yselect1 = yselect[list(range(bkg_start,len(yselect)))]

        k = 25
        slope = None
        while k < 100:
            slope = yselect1[k]-yselect1[k-1]
            if slope < bkg_slope*1.5:
                k = k + 1          
            else:
                bkg_end = k
                k = 100

        #fitting the background with a linear model
        x_bkg = xselect1[list(range(0,bkg_end))]
        y_bkg = yselect1[list(range(0,bkg_end))]    
        out_bkg = mod.fit(y_bkg, x = x_bkg)
        bkg_x_intercept = (-out_bkg.params['reg_intercept'].value/out_bkg.params['reg_slope'].value)

        if fit_background == True:
            #finding the valence onset

            xselect_valence = x_reversed[list(range(bkg_end+bkg_start+startvalue,len(x)))]
            yselect_valence = y_reversed[list(range(bkg_end+bkg_start+startvalue,len(y)))]

            k = 5
            slope = None
            while k < len(yselect_valence)-1:
                slope = yselect_valence[k]-yselect_valence[k-1]
                if slope < bkg_slope*4:
                    k = k + 1
                else:
                    valence_start = k
                    valence_slope = slope
                    k = len(yselect_valence)-1

            #finding valence end
            xselect_valence1 = xselect_valence[list(range(valence_start,len(xselect_valence)))]
            yselect_valence1 = yselect_valence[list(range(valence_start,len(yselect_valence)))]

            k = 15
            slope = None
            while k < len(yselect_valence1)-1:
                slope = yselect_valence1[k]-yselect_valence1[k-1]
                if slope < valence_slope*1.2:
                    k = k + 1          
                else:
                    valence_end = k
                    k = len(yselect_valence1)-1

            x_valence = xselect_valence1[list(range(0,valence_end))]
            y_valence = yselect_valence1[list(range(0,valence_end))]
            out_valence = mod.fit(y_valence, x = x_valence)

            fits_intercept = ((out_valence.params['reg_intercept'].value-out_bkg.params['reg_intercept'].value)/(out_bkg.params['reg_slope'].value-out_valence.params['reg_slope'].value))

            if withplots == 1:
                plt.plot(x,y, label = 'data')
                plt.plot(x_valence,out_valence.best_fit, label = 'valence fit')
                plt.plot(xselect_valence[valence_start],yselect_valence[valence_start],'o', label = 'valence_start')
                plt.plot(xselect_valence1[valence_end],yselect_valence1[valence_end],'o', label = 'valence_end')
                plt.plot(xselect[bkg_start], yselect[bkg_start], 'o',label = 'bkg_start')
                plt.plot(xselect1[bkg_end], yselect1[bkg_end], 'o',label = 'bkg_end')
                plt.plot(x_bkg, out_bkg.best_fit, label='background fit')
                plt.xlabel(col_BE[i][1])
                plt.ylabel(col_counts[i][1])
                plt.title(col_counts[i][0])
                plt.yscale(plotscale)
                plt.xlim(-4, 8)
                plt.ylim(-1e4,2e6)
                plt.legend()
                plt.show()
                
                print("bkg_x_intercept:\n",bkg_x_intercept)
                print("fits_intercept:\n",fits_intercept)

            

            intercepts = np.vstack((bkg_x_intercept,fits_intercept)).T
            UPS_header = pd.MultiIndex.from_product([[col_BE[i][0]],['bkg_x_intercept','fits_intercept']],names=['Coordinate','Data type'])
            UPS_output = pd.DataFrame(data=intercepts, columns=UPS_header)

        else:
            if withplots == 1:
                plt.plot(x,y, label = 'data')
                plt.plot(x_bkg, out_bkg.best_fit, label='valence fit')
                plt.xlabel(col_BE[i][1])
                plt.ylabel(col_counts[i][1])
                plt.title(col_counts[i][0])
                plt.yscale(plotscale)
                plt.xlim(0, 8)
                plt.ylim(-1e4,2e6)
                plt.legend()
                plt.show()
                
                print("bkg_x_intercept:\n",bkg_x_intercept)

            UPS_header = pd.MultiIndex.from_product([[col_BE[i][0]],['bkg_x_intercept']],names=['Coordinate','Data type'])
            UPS_output = pd.DataFrame(data=np.array([bkg_x_intercept]), columns=UPS_header)

        UPS_outframe = pd.concat([UPS_outframe, UPS_output], axis = 1)
    return UPS_outframe

##########################
# Functions related to REELS analysis
##########################
def REELS_fit(data, withplots = True,plotscale = 'linear'):
    '''Fit REELS data using dataframe input from read_REELS, the data is fitted by finding the the 0 energy loss peak position, and the onset of the energy loss curve
    Output: dataframe with the onset and peak positions, as well as the band gap and the raw data.'''
    column_headers = data.columns.values
    column_headers = data.columns.values
    col_BE = column_headers[::2]
    col_counts = column_headers[1::2]
    mod = LinearModel(prefix='reg_')


    BGS = np.array([])
    REELS_outframe = pd.DataFrame()
    REELS_data = pd.DataFrame()
    for i in range(0,len(col_BE)):
        x = data[col_BE[i]]
        y = data[col_counts[i]]
        
        #finds the 0 energy loss peak
        peaks, _ = find_peaks(y,prominence= 1000000)

        #selects the data after the peak
        xselect1 = x[list(range(0,peaks[0]))].values
        yselect1 = y[list(range(0,peaks[0]))].values
        
        #formating
        yselect1 = yselect1[::-1]
        xselect1 = xselect1[::-1]
        peak = x[peaks].values[0]

        #locates the onset of the energy loss.
        j = False
        k = 0
        while j != True:
            slope = yselect1[k+1]-yselect1[k]
            if slope < 1400:
                k = k + 1
            else:
                idx2 = k
                j = True
            if k == 100:
                j = True                   

        
        onset = xselect1[idx2]
        #finds the onset from the peak location and onset
        BG = onset-peak

        if withplots == 1:
            plt.plot(x,y, label = 'data')
            plt.plot(x[peaks], y[peaks], 'o', label = 'elastic scattering peak')
            plt.plot(xselect1[idx2], yselect1[idx2], 'o', label = 'onset of energy loss')
            plt.xlabel(col_BE[i][1])
            plt.ylabel(col_counts[i][1])
            plt.title(col_counts[i][0])
            plt.yscale(plotscale)
            plt.xlim(20, 4)
            plt.legend()
            plt.show()

            #print output peak positions, intensity, and FWHM
            print("Onset:\n",onset)
            print("Peak:\n",peak)
            print("Band gap:\n",BG)

        #constructs output dataframe of band gap, onset and Peak values.
        BGS = np.vstack((BG,xselect1[idx2],x[peaks])).T
        REELS_header = pd.MultiIndex.from_product([[col_BE[i][0]],['Band gap','Onset BE', 'Peak BE']],names=['Coordinate','Data type'])
        REELS_bandgaps = pd.DataFrame(data=BGS, columns=REELS_header)       

        #construct output for the raw data
        raw = np.vstack((x,y)).T
        raw_header = pd.MultiIndex.from_product([[col_BE[i][0]],['BE (ev)','Intensity (counts)']],names=['Coordinate','Data type'])
        REELS_data = pd.DataFrame(data=raw, columns=raw_header)

        #assembles the final output
        REELS_outframe = pd.concat([REELS_outframe,REELS_data, REELS_bandgaps], axis = 1)
    return REELS_outframe


##########################
# Functions related to plotting
##########################

def plot_grid(coords, grid):
    '''"Plot a set of real measurement points on a custom grid defined with the "measurement_grid" function. The corrected grid locations are shown."'''
    corrected_grid = coords_to_grid(coords, grid)
    plt.scatter(grid.iloc[:,0], grid.iloc[:,1], color = 'black', s = 80)
    plt.scatter(coords.iloc[:,0], coords.iloc[:,1], color = 'green', s = 20)
    plt.scatter(corrected_grid.iloc[:,0], corrected_grid.iloc[:,1], color = 'red', s = 20)
    plt.legend(['Defined grid', 'Measured', 'Corrected'])


def plot_data(data, datatype_x, datatype_y, x = "all", y = "all",datatype_select = None,datatype_select_value = None, legend = True, scatter_plot = False,plotscale = "linear", title = "auto"):
    '''Creates a XY plot/scatter plot based on datatype from a dataframe'''

    #x and y to list if only 1 value specified
    if type(x) != list:
        x = [x]
    if type(y) != list:
        y = [y]
    x_data = []
    y_data = []
    labels = []
    #extracts the specified data point by point
    for i in range(len(x)):
        x_data.append(get_data(data, datatype_x, x[i], y[i], False,False))
        y_data.append(get_data(data, datatype_y, x[i], y[i], False,False))
        if x[0] == "all" and y[0] == "all":
            labels = data.columns.get_level_values(0).unique().values
        else:
            grid = MI_to_grid(data)
            xcoord, ycoord = closest_coord(grid, x[i], y[i])
            labels.append('{:.1f},{:.1f}'.format(xcoord, ycoord))
    
    colors = plt.cm.jet(np.linspace(0, 1, len(labels))) #data.columns.get_level_values(0).unique().values

    #formating
    if len(labels) == 1:
        labels = labels[0]
    if x[0] == "all" and y[0] == "all":
        x_data = x_data[0]
        y_data = y_data[0]
    else:
        x_data = np.transpose(x_data)
        y_data = np.transpose(y_data)
    
    #if datatype with multiple values per point is plotted only selects one value, based on the datatype_select, datatype_select_value. 
    if datatype_select != None:
        y_data = y_data.iloc[data.index[data[data.iloc[:, data.columns.get_level_values(1)== datatype_select].columns[0]] == datatype_select_value]]
        x_data_coords = x_data.columns.get_level_values(0)
        y_data_coords = y_data.columns.get_level_values(0)
        data_coords = [j for j in x_data_coords if j not in y_data_coords]
        x_data.drop(data_coords, level=0,axis = 1, inplace=True) 
        x_data = x_data.values[0]
        y_data = y_data.values[0]
        labels = datatype_select + ': ' + str(round(datatype_select_value,2))

#plots scatter plot if scatter_plot is not false, else line plot
    if x[0] == "all" and y[0] == "all":
        for idx, (x_val, y_val) in enumerate(zip(x_data.values.T, y_data.values.T)):
            if scatter_plot:
                plt.plot(x_val, y_val, 'o', color=colors[idx], label=labels[idx])
            else:
                plt.plot(x_val, y_val, color=colors[idx], label=labels[idx])
    else:
        for idx, (x_val, y_val) in enumerate(zip(x_data.T, y_data.T)):
            if scatter_plot:
                plt.plot(x_val, y_val, 'o', color=colors[idx], label=labels[idx])
            else:
                plt.plot(x_val, y_val, color=colors[idx], label=labels[idx])
    plt.xlabel(datatype_x)
    plt.ylabel(datatype_y)
    plt.yscale(plotscale)
    if legend == True:
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    if title == "auto":
        plt.title("{} over {}".format(datatype_y, datatype_x))
    else:
        plt.title(title)

def heatmap(data, datatype, title = None, datatype_select = None, datatype_select_value = None, min_limit = None, max_limit = None, excluded_x = None, excluded_y = None):
    '''Creates XY heatmap based on datatype.'''
    xy = MI_to_grid(data).drop_duplicates(ignore_index=True)

    #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    # I have commented this line out as it broke when nans from divide by zeros from math_on_columns were in dataframe
    # Hopefully nothing else breaks because of it? (Tested and found no problems at least...)
    #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
    # remove columns containing only nan values
    #data = data.dropna(axis = 1, how = 'all')
    

    if datatype_select_value == None or datatype_select == None:
        values = data.iloc[:, data.columns.get_level_values(1)==datatype]
        values = values.dropna(axis = 0, how = 'all').values
    else:
        # find indices of closest values to data_select_value
        datatype_select_value_close_index = np.abs(data.iloc[:, data.columns.get_level_values(1)==datatype_select] - datatype_select_value).idxmin()
        # select datatype values
        values = []
        for i in range(len(datatype_select_value_close_index)):
            datatype_select_value_close = data.iloc[:, data.columns.get_level_values(1)==datatype_select].iloc[datatype_select_value_close_index[i],i]
            close_value = data.iloc[:, data.columns.get_level_values(1)==datatype].iloc[data.index[data[data.iloc[:, data.columns.get_level_values(1)== datatype_select].columns[i]] == datatype_select_value_close]].iloc[0,i]
            values.append(close_value)
        
    dfPlot = np.vstack([np.round(xy['x'],1),np.round(xy['y'],1),values]).T
    dfPlot = dfPlot.astype('float')
    dfPlot = pd.DataFrame(dfPlot, columns = ['X (mm)','Y (mm)','Data'])

    # fill in excluded data with a color instead of leaving it transparent
    # limit to min and max range
    tablelist = []
    if min_limit != None:
        tablelist.append([dfPlot['Data'] < min_limit][0])
    if max_limit != None:
        tablelist.append([dfPlot['Data'] > max_limit][0])
    # also use manually chosen coordinates to exclude
    if excluded_x != None:
        # convert coordinate to list (if only a single point is provided)
        if type(excluded_x) != list:
            excluded_x = [excluded_x]
        if type(excluded_y) != list:
            excluded_y = [excluded_y]
        # find the grid the data is using
        grid = pd.concat([dfPlot['X (mm)'],dfPlot['Y (mm)']], axis = 1)
        grid.rename(columns={'X (mm)':'x','Y (mm)':'y'}, inplace = True)
        # make tables of chosen xy coordinates
        for i in range(len(excluded_x)):
            coords = closest_coord(grid, excluded_x[i], excluded_y[i])
            xtrue = dfPlot['X (mm)'] == coords[0]
            ytrue = dfPlot['Y (mm)'] == coords[1]
            xytrue = [xtrue[i] and ytrue[i] for i in range(len(xtrue))]
            tablelist.append(xytrue)

    plotnans = 0
    if min_limit or max_limit or excluded_x != None:
        # sum up all excluded values in a final exclusion table
        table = np.array([sum(i) for i in zip(*tablelist)]).astype(bool)

        dfPlot_nans = dfPlot.copy()
        dfPlot_nans['Data'].loc[~table] = float('NaN')
        dfPlot_nans['Data'].loc[dfPlot_nans['Data'].isna() == False] = 100

        dfPlot_nans = dfPlot_nans.pivot(index='Y (mm)', columns='X (mm)', values='Data')
        dfPlot_nans = dfPlot_nans.iloc[::-1]
        plotnans = 1

    dfPlot_vals = dfPlot.copy()
    if min_limit or max_limit or excluded_x != None:
        dfPlot_vals['Data'].loc[table] = float('NaN')
    dfPlot_vals = dfPlot_vals.pivot(index='Y (mm)', columns='X (mm)', values='Data')
    dfPlot_vals = dfPlot_vals.iloc[::-1]
    sns.heatmap(dfPlot_vals, annot=False, cmap='viridis', fmt=".1f")#, xticklabels=dfPlot_vals.columns.values, yticklabels=dfPlot_vals.index.values)
    if plotnans == 1:
        sns.heatmap(dfPlot_nans, annot=False, fmt=".4f", cbar = False, cmap = "mako")
    plt.title(title)

def new_heatmap(datatype, data=None, filepath = None, exclude=None, savepath=None, title=None):
    "plot heatmaps with interpolated background, like in Nomad, if savepath ends with .png, it will save as png, if it ends with .html, it will save as html (interactive)"

    if filepath is not None:
        raw_data = pd.read_excel(filepath, header=0)
        x = raw_data["X (mm)"].values
        y = raw_data["Y (mm)"].values
        z = raw_data[datatype].values

    if data is not None: 
        if exclude != None:
            for point in exclude:
                data = data.drop(data.iloc[:,data.columns.get_level_values(0)==point], axis=1)
        xy = MI_to_grid(data).drop_duplicates(ignore_index=True)
        x = xy["x"].values
        y = xy["y"].values
        z = data.iloc[:, data.columns.get_level_values(1)==datatype].dropna().values.flatten()

    xi = np.linspace(min(x), max(x), 100)
    yi = np.linspace(min(y), max(y), 100)
    xi, yi = np.meshgrid(xi, yi)
    zi = griddata((x, y), z, (xi, yi), method='linear')

    scatter = go.Scatter(
            x=x,
            y=y,
            mode='markers',
            marker=dict(
                size=15,
                color=z,  # Set color to thickness values
                colorscale='Viridis',  # Choose a colorscale
                showscale=False,  # Hide the colorbar for the scatter plot
                line=dict(
                    width=2,  # Set the width of the border
                    color='DarkSlateGrey'  # Set the color of the border
            ) ),
        )
    if datatype == "Layer 1 Thickness (nm)":
        cbar_title = "Thickness (nm)"
    elif datatype == "Layer 1 P Atomic %":
        cbar_title = "P Atomic %"
    elif datatype == "Layer 1 S Atomic %":
        cbar_title = "S Atomic %"
    elif datatype == "Layer 1 Cu Atomic %":
        cbar_title = "Cu Atomic %"
    elif datatype == "Layer 1 Zr Atomic %":
        cbar_title = "Zr Atomic %"
    elif datatype == "Layer 1 Ba Atomic %":
        cbar_title = "Ba Atomic %"
    else:
        cbar_title = datatype

    heatmap = go.Heatmap(
    x=xi[0],
    y=yi[:, 0],
    z=zi,
    colorscale='Viridis',
    colorbar=dict(title=cbar_title),
    #zmin = 10, zmax = 60
    )

    fig = go.Figure(data=[heatmap, scatter])

    if title == None:
        title = datatype
    
    fig.update_layout(title=title,
    xaxis_title='X Position (mm)',
    yaxis_title='Y Position (mm)',
    template='plotly_white',
    autosize =False,
    width = 600,
    height = 500)

    if savepath:
        if savepath.endswith(".png"):
            fig.write_image(savepath, scale=2) #scale sets the resolution here
        
        if savepath.endswith(".html"):
            fig.write_html(savepath)
    
    fig.show()

def lp_translate_excel(folder,filename):
    """Creates a new excel file with translated coordinates, given the coordinates 
    of the corners in Sheet2, assuming they are stored rightafter the statistics"""
    filepath =os.path.join(folder,filename+".xlsx")
    new_path = os.path.join(folder,filename+"_translated.xlsx")

    first_data = pd.read_excel(filepath, sheet_name = "Sheet1")

    first_x = first_data["X (mm)"]
    first_y = first_data["Y (mm)"]

    corners = pd.read_excel(filepath, sheet_name = "Sheet2", usecols=(6,7))

    trans_x = corners.iloc[[0,1],0].mean()
    trans_y = corners.iloc[[0,1],1].mean()

    new_x = first_x - trans_x
    new_y = first_y - trans_y

    new_data = first_data.copy()
    new_data["X (mm)"] = new_x
    new_data["Y (mm)"] = new_y

    new_data.to_excel(new_path, index = False)

def plot_scatter_colormap(data, datatype_x, datatype_y, datatype_z, x = "all", y = "all",datatype_select = None,datatype_select_value = None,min_limit = None, max_limit = None,plotscale = "linear", title = "auto",colormap_label = None):
    '''Creates a XY plot/scatter plot based on datatype'''
    #x and y to list if only 1 value specified
    if type(x) != list:
        x = [x]
    if type(y) != list:
        y = [y]
    x_data = []
    y_data = []
    z_data = []
    labels = []
    #extracts the specified data point by point
    for i in range(len(x)):
        x_data.append(get_data(data, datatype_x, x[i], y[i], False,False))
        y_data.append(get_data(data, datatype_y, x[i], y[i], False,False))
        z_data.append(get_data(data, datatype_z, x[i], y[i], False,False))
        if x[0] == "all" and y[0] == "all":
            labels = data.columns.get_level_values(0).unique().values
            
        else:
            grid = MI_to_grid(data)
            xcoord, ycoord = closest_coord(grid, x[i], y[i])
            labels.append('{:.1f},{:.1f}'.format(xcoord, ycoord))

    #formating
    if len(labels) == 1:
        labels = labels[0]
    if x[0] == "all" and y[0] == "all":
        x_data = x_data[0]
        y_data = y_data[0]
        z_data = z_data[0]
    else:
        x_data = np.transpose(x_data)
        y_data = np.transpose(y_data)
        z_data = np.transpose(z_data)

    #if datatype with multiple values per point is plotted only selects one value, based on the datatype_select, datatype_select_value. 
    if datatype_select != None:
        z_data = z_data.iloc[data.index[data[data.iloc[:, data.columns.get_level_values(1)== datatype_select].columns[0]] == datatype_select_value]]
        x_data_coords = x_data.columns.get_level_values(0)
        z_data_coords = z_data.columns.get_level_values(0)
        data_coords = [j for j in x_data_coords if j not in z_data_coords]
        x_data.drop(data_coords, level=0,axis = 1, inplace=True)
        y_data.drop(data_coords, level=0,axis = 1, inplace=True) 
        x_data = x_data.values[0]
        y_data = y_data.values[0]
        z_data = z_data.values[0]

    #removes data points above the max limit of the z data
    if max_limit != None:
        x_data = x_data[z_data<max_limit]
        y_data = y_data[z_data<max_limit]
        z_data = z_data[z_data<max_limit]

    #removes data points below the min limit of the z data
    if min_limit != None:
        x_data = x_data[z_data>min_limit]
        y_data = y_data[z_data>min_limit]
        z_data = z_data[z_data>min_limit]

    #Defines the color from the z_data
    colors = z_data
    plt.scatter(x_data, y_data,c = colors ,cmap = 'viridis')
    plt.colorbar().set_label(colormap_label)
    plt.xlabel(datatype_x)
    plt.ylabel(datatype_y)
    plt.yscale(plotscale)
    if title == "auto":
        plt.title("{} over {}".format(datatype_y, datatype_x))
    else:
        plt.title(title)

##########################
# Functions related to combining, examining, saving, and loading data
##########################

def closest_coord(grid, x, y):
    '''"Find closest x and y coordinate for a grid."'''
    xminindex = np.abs(grid - x).idxmin().iloc[0]
    xcoord = grid.iloc[xminindex, 0]
    yminindex = np.abs(grid[grid['x']==xcoord] - y).idxmin().iloc[1]
    ycoord = grid.iloc[yminindex, 1]
    return xcoord, ycoord

def combine_data(datalist):
    '''"Combine multiple measurements into a single dataframe."'''
    dataframe = pd.concat(datalist, axis=1)
    return dataframe

def compare_atomic_XPS(data, type1, type2):
    '''"Outdated, use math_on_columns instead."'''
    row = []
    datalength = len(data.columns.get_level_values(1).unique())
    for i in range(0, len(data.values[0]), datalength):
        atomic1 = data.loc[data.iloc[:,i] == type1].iloc[:,i+datalength-1]
        atomic2 = data.loc[data.iloc[:,i] == type2].iloc[:,i+datalength-1]
        compareatomic = round(atomic1.iloc[0] / atomic2.iloc[0], 3)
        rowname = "{}/{}".format(type1.split()[0], type2.split()[0])
        # append row
        row.append([rowname, float('NaN'), float('NaN'), float('NaN'), float('NaN'), compareatomic])

    datarows = np.vstack([data.values, np.concatenate(row)])
    coord_grid = MI_to_grid(data).drop_duplicates(ignore_index = True)
    coord_header = grid_to_MIheader(coord_grid)
    # construct dataframe with multiindexing for coordinates
    header = pd.MultiIndex.from_product([coord_header, data.columns.get_level_values(1).unique()],names=['Coordinate','Data type'])
    data = pd.DataFrame(datarows, columns=header)
    data.iloc[-1,:].replace('nan', float("NaN"), inplace = True)
    return data

def math_on_columns(data, type1, type2, operation = "/"):
    '''"Perform an operation on two columns in a provided dataframe. Usage: math_on_columns(data, datatype1, datatype2, operation), where "operation" can be +, -, *, or /."'''
    coordinatelength = len(data.columns.get_level_values(0).unique())
    headerlength = len(data.columns.get_level_values(1).unique())
    k = 0
    # do math on values
    data = data.copy()
    for i in range(coordinatelength):
        val1 = data.iloc[:, data.columns.get_level_values(1)==type1].iloc[:,i]
        if isinstance(type2, str):
            val2 = data.iloc[:, data.columns.get_level_values(1)==type2].iloc[:,i]
        if isinstance(type2, (int,float)):
            val2= type2
        if operation == "+":
            resultval = val1 + val2
        elif operation == "-":
            resultval = val1 - val2
        elif operation == "*":
            resultval = val1 * val2
        elif operation == "/":
            try:
                resultval= val1 / val2
            except ZeroDivisionError:
                resultval = float("NaN")
        # insert result
        data.insert(headerlength*(i+1)+k, "{}".format(data.columns.get_level_values(0).unique()[i]), resultval, allow_duplicates=True)
        k += 1

    # rename added columns
    if operation == "+":
        rowname = "{} + {}".format(type1, type2)
    elif operation == "-":
        rowname = "{} - {}".format(type1, type2)
    elif operation == "*":
        rowname = "{} * {}".format(type1, type2)
    elif operation == "/":
        rowname = "{} / {}".format(type1, type2)
    data.rename(columns={'':rowname}, inplace = True)
    return data

def get_data(data, type = 'all', x = 'all', y = 'all', printinfo = True, drop_nan = True):
    '''"Print a data type from a multi index dataframe at a specific coordinate. The coordinate does not have to be exact. Leave type as blank or 'all' to select all types. Leave coordinates blank or 'all' to select all coordinates."'''
    if x == 'all' and y == 'all':
        if type == 'all':
            if printinfo == True:
                print("All data at all coordinates.")
            if drop_nan == True:
                data = data.dropna(axis = 0, how = 'all').fillna('-')
            return data
        else:
            if printinfo == True:
                print("{} data at all coordinates.".format(type))
            if drop_nan == True:
                data = data.dropna(axis = 0, how = 'all').fillna('-')
            return data.iloc[:, data.columns.get_level_values(1)==type]
    else:
        datagrid = MI_to_grid(data)
        # find closest x and y coordinate
        xcoord, ycoord = closest_coord(datagrid, x, y)
        coords = '{},{}'.format(xcoord, ycoord)
        if type == 'all':
            if printinfo == True:
                print("All data at {},{}.".format(x, y))
            if drop_nan == True:
                data = data.dropna(axis = 0, how = 'all').fillna('-')
            return data.xs(coords, axis=1)
        else:
            if printinfo == True:
                print("{} data at {},{}.".format(type, x, y))
            if drop_nan == True:
                data = data.dropna(axis = 0, how = 'all').fillna('-')
            return data.xs(coords, axis=1)[type]
        
def translate_data(data, x, y):
    '''"Move a set of datapoints by a given x and y offset. Useful when combining multiple samples into one dataframe."'''
    coords = MI_to_grid(data)
    coords['x'] = coords['x'] + x
    coords['y'] = coords['y'] + y
    coord_header = grid_to_MIheader(coords)
    header = pd.MultiIndex.from_arrays([coord_header, data.columns.get_level_values(1)],names=['Coordinate','Data type'])
    data = pd.DataFrame(data.values, columns=header)
    return data, coords

def get_data_index(data, x, y):
    '''"Recommended to use get_data instead. Does not work for XRD data. Prints data from a multi index dataframe at a specific index."'''
    datagrid = MI_to_grid(data)
    xlist = datagrid['x'].unique()
    ylist = datagrid['y'].unique()
    coords = '{},{}'.format(xlist[x], ylist[y])
    print("Data at index {},{} ({}):".format(x, y, coords))
    return data.xs(coords, axis=1)

def save_data(dataframe, filename, separator = "\t"):
    '''"Save dataframe to tab seperated txt file."'''
    dataframe.to_csv(filename, separator, index=False, encoding='utf-8')
    return

def load_data(filepath, separator = "\t"):
    '''"Load txt to dataframe."'''
    dataframe = pd.read_csv(filepath, sep=separator, header=[0, 1])
    dataframe.columns.rename(["Coordinate", "Data type"], level=[0, 1], inplace = True)
    return dataframe

def rename_SE_images(folderpath):
    # Define the directory containing the images
    directory = folderpath  # Replace with the path to your folder

    # Function to extract the number from the filename
    def extract_number(filename):
        match = re.search(r'Electron Image (\d+)\.bmp', filename)
        return int(match.group(1)) if match else None

    # Get a list of all .bmp files in the directory
    files = [f for f in os.listdir(directory) if f.endswith('.bmp')]

    # Sort the files by their extracted number
    files.sort(key=extract_number)

    # Rename the files sequentially
    for i, filename in enumerate(files, start=1):
        new_name = f'Electron Image {i}.bmp'
        old_file_path = os.path.join(directory, filename)
        new_file_path = os.path.join(directory, new_name)
        
        os.rename(old_file_path, new_file_path)
        print(f'Renamed: {filename} -> {new_name}')

    print("Renaming completed.")

def old_EDS_coordinates(ncolumns, nrows, mag, spacing, filepath, new_path, edge=4,rotate=False):

    if rotate == "90":
        ncolumns, nrows = nrows, ncolumns
        areax = 2.8 * 100 / mag
        areay = 4.1 * 100 / mag
    
    if rotate ==False:
        # Calculate the effective area size in the x-direction, considering magnification, 
        # assuming the x/y ratio is constant 4.1 : 2.8
        areax = 4.1 * 100 / mag
        areay= 2.8*100 / mag
    
    # Calculate the spacing , gridlength and starting x-coordinate for the grid in x-direction (assuming the grid is centered)
    space_x = areax * spacing / 100
    gridlength = (ncolumns - 1) * (space_x + areax) + areax
    startx = -gridlength / 2 + (areax / 2)

    # do the same for the y-direction
    
    space_y = areay*spacing/100
    gridheight= (nrows-1)*(space_y+ areay) + areay
    starty = -gridheight/2 + (areay/2)

    # Check if the grid dimensions exceed the maximum allowed size (31x31 mm)
    # if so, reduce the spacing by 10% and try again 
    if gridlength >= 39-edge or gridheight >= 39-edge:
        print("Spacing is too large for the map")

        new_spacing = np.round(spacing - spacing * 0.1, 0)
        print("New spacing is", new_spacing)

        return EDS_coordinates(ncolumns, nrows, mag, new_spacing, filepath, new_path,rotate)

    # Create a list to hold grid parameters (input of the grid function)
    grid_input = [ncolumns, nrows,np.round(-startx*2, 2), np.round(-starty*2, 2), np.round(startx, 2), np.round(starty, 2)]

    # Generate coordinates for each column
    coord_x = np.round(np.linspace(startx, -startx, ncolumns), 2)
    coord_y = np.round(np.linspace(starty, -starty, nrows), 2)
    X=[]
    Y=[]
    for j in range(0, ncolumns):
        for i in range(0, nrows):
            Y.append(coord_y[i])
            X.append(coord_x[j])

    first_data = pd.read_excel(filepath, sheet_name = "Sheet1")

    new_data = first_data.copy()
    new_data["X (mm)"] = X
    new_data["Y (mm)"] = Y

    new_data.to_excel(new_path, index = False)

    return X,Y, grid_input, areax, areay

def read_CRAIC(file_path, header_lines=10, print_header= True):
    header =[]
    with open(file_path, "r") as file:
        for _ in range(header_lines):
            #file.readline()
            header.append(file.readline().strip())
        wavelengths_line = file.readline().strip().split("\t")
        intensities_line = file.readline().strip().split("\t")
    wavelengths = [float(w) for w in wavelengths_line]
    intensities = [float(i) for i in intensities_line]

    data = pd.DataFrame({"Wavelength": wavelengths, "Intensity": intensities})

    if print_header==True:
        print("Header Lines:")

        for line in header:
            print(line)
    return data

def snake_grid(x, y): # x and y are lists of coordinates you should take note of 
    X_snake = []
    Y_snake = []

    # Loop through each y-coordinate from bottom to top
    for i, y_val in enumerate(y):

        if i % 2 == 0: 
            X_snake.extend(x) # Even row: left to right ( add x normally)
        else: 
            X_snake.extend(x[::-1]) # Odd row: right to left (add x in reverse)
        Y_snake.extend([y_val] * len(x)) # add as many y values as x values

    grid_snake = pd.DataFrame({"x": X_snake, "y": Y_snake})
    return grid_snake
    
def CRAIC_map(folder, background, reflection_name, transmission_name, grid, unit= "nm", with_plots=True, savepath=None):
    # x axis is taken from the background file, maybe check that it is always the same
    data = pd.DataFrame()
    npoints = len(grid)
    background = read_CRAIC(os.path.join(folder, background), print_header=False)
    fig, ax = plt.subplots(3,1, figsize=(10, 10))

    if unit == "nm":
        x_axis = background["Wavelength"]
        x_label = "Wavelength (nm)"

    elif unit == "eV":
        h = 4.135*10**-15
        c = 3*10**8
        wavelength_ev = (h*c)/(background["Wavelength"]*10**-9)

        x_axis = wavelength_ev
        x_label = "Energy (eV)"

    for i in range(1, npoints+1):
            
        file_refl = f"{reflection_name}-{i}.msp"
        file_transl = f"{transmission_name}-{i}.msp"

        data_R = read_CRAIC(os.path.join(folder, file_refl), print_header=False)
        data_T = read_CRAIC(os.path.join(folder, file_transl), print_header=False)

        data_R["Intensity"] = data_R["Intensity"] - background["Intensity"]
        #calculate absorption coefficient
        data_A = -(np.log(data_T["Intensity"]/(100-data_R["Intensity"])))*10**5 

        if with_plots == True:
            ax[0].plot(x_axis, data_R["Intensity"], label=f" {i}") 
            ax[1].plot(x_axis, data_T["Intensity"], label=f" {i}")
            ax[2].plot(x_axis, data_A, label=f" {i}")

            ax[0].legend(bbox_to_anchor=(1, 1), loc='upper right')
            ax[1].legend(bbox_to_anchor=(1, 1), loc='upper right')
            ax[2].legend(bbox_to_anchor=(1, 1), loc='upper right')

            ax[0].set_ylabel("R %")
            ax[1].set_ylabel("T %")
            ax[2].set_ylabel(r"$\alpha$ (cm$^{-1}$)")

            for ax_ in ax:
                ax_.set_xlabel(x_label)
            
        plt.tight_layout()
        if savepath: 
            plt.savefig(savepath, dpi=300)
    
        # save data in a multiindex dataframe

        data_R["Intensity"].rename(f"R", inplace=True)
        data_T["Intensity"].rename(f"T", inplace=True)
        data_A.rename(f"A", inplace=True)
        data = pd.concat([data, data_R["Wavelength"]], axis=1, ignore_index=False)
        data = pd.concat([data, data_R["Intensity"]], axis=1, ignore_index=False)
        data = pd.concat([data, data_T["Intensity"]], axis=1, ignore_index=False)
        data = pd.concat([data, data_A], axis=1, ignore_index=False)

    coord_header = grid_to_MIheader(grid) # check how the points are collected and build the grid accordingly
    df_header = pd.MultiIndex.from_product([coord_header, data.columns[0:4]],names=['Coordinate','Data type'])
    data_MI = pd.DataFrame(data.values, columns=df_header)

    return data_MI


def plot_XRD_shift_subplots(data, datatype_x, datatype_y, x, y_list, shift, title, material_guess, nrows, ncols, figsize=(12, 10), save=True):
    """
    Plots XRD shift for multiple y-coordinates in subplots.

    Parameters:
    - data: DataFrame containing the data
    - datatype_x: str, type of data for x-axis
    - datatype_y: str, type of data for y-axis
    - x: list, list of x-coordinates
    - y_list: list of lists, each sublist contains y-coordinates for a subplot
    - shift: float, value by which to shift the y-axis data
    - title: str, title of the entire figure
    - ref_lines: array, reference lines to be plotted
    - nrows: int, number of rows of subplots
    - ncols: int, number of columns of subplots
    - figsize: tuple, size of the figure
    - save: bool, whether to save the figure as a file

    Returns:
    - plots the XRD data for multiple y-coordinates in subplots
    """
    with open (os.path.join("XRD", "reflections", "reflections.pkl"), "rb") as file:
        ref_peaks_df = pickle.load(file)

    ref_peaks = ref_peaks_df[material_guess]
    ref_lines = ref_peaks["Peak 2theta"][ref_peaks["Peak 2theta"].notna()].values
    fig, axes = plt.subplots(nrows=nrows, ncols=ncols, figsize=figsize)
    axes = axes.flatten()

    for idx, pos in enumerate(y_list):
        ax = axes[idx]
        for i in range(len(x)):
            #print('x =', x[i], 'y =', pos[i])
            x_data = get_data(data, datatype_x, x[i], pos[i], printinfo=False, drop_nan=False)
            y_data = get_data(data, datatype_y, x[i], pos[i], printinfo=False, drop_nan=False)
            lab = "{:.1f},{:.1f}".format(x[i], pos[i])

            ax.plot(x_data, y_data + shift * i, label=lab)

        ax.set_title(f'Y = {pos[0]}')
        ax.legend(bbox_to_anchor=(1.01, 1), loc='upper left')

        if ref_lines is not None:
            for line in ref_lines:
                ax.axvline(x=line, linestyle='--', alpha=0.5, color='grey')


    axes[-1].plot(ref_peaks["2theta"], ref_peaks["I"], label=str(material_guess)) 
    #axes[-1].axvline(x=ref_lines.values, linestyle='--', alpha=0.5, color='grey')      
    axes[-1].legend(bbox_to_anchor=(1.01, 1), loc='upper left')
    plt.suptitle(title)
    plt.tight_layout(rect=[0, 0, 1, 0.97])

    if save:
        plt.savefig(f'{title}_XRD_shift_subplots.png', dpi=120, bbox_inches='tight')

    plt.show()


def plot_XRD_shift(data,datatype_x, datatype_y,  shift,x,y, title=None, savepath= False): #x, y = list of points to plot]
    x_data = []
    y_data = []
    labels = []
    plt.figure(figsize = (12,5))
    for i in range(len(x)):
        x_data.append(get_data(data, datatype_x, x[i], y[i], False,False))
        y_data.append(get_data(data, datatype_y, x[i], y[i], False,False))
        if x[0] == "all" and y[0] == "all":
            labels = data.columns.get_level_values(0).unique().values
        else:
            grid = MI_to_grid(data)
            xcoord, ycoord = closest_coord(grid, x[i], y[i])
            labels.append('{:.1f},{:.1f}'.format(xcoord, ycoord))

        plt.plot(x_data[i], y_data[i]+ shift*i, label = labels[i])
    plt.xlabel(datatype_x)
    plt.ylabel(datatype_y)
    plt.title(title)
    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    
    if savepath:
        path = os.path.join('plots', title + 'shift.png')
        plt.savefig(path, dpi=120, bbox_inches='tight')
        
    plt.show()

def fit_two_related_peaks(x, y):

    # Initialize two Pseudo-Voigt models with prefixes to distinguish parameters
    model1 = PseudoVoigtModel(prefix='p1_')
    model2 = PseudoVoigtModel(prefix='p2_')

    # Estimate initial parameters for the first peak
    params = model1.guess(y, x=x)
    
    # Extract initial guesses
    amplitude1 = params['p1_amplitude'].value
    center1 = params['p1_center'].value
    sigma1 = params['p1_sigma'].value
    fraction1 = params['p1_fraction'].value

    # Set constraints for the second peak based on the provided relations
    #xpeak2 = 2 * np.arcsin((0.154439 / 0.1540562) * np.sin(center1 / 2))
    xpeak2= (360/np.pi)* np.arcsin((0.154439 / 0.1540562) * np.sin(center1*np.pi /360))
    
    params.add('p2_center', expr='(360/pi)* arcsin((0.154439 / 0.1540562) * sin(p1_center*pi /360))')
    params.add('p2_amplitude', expr='0.5 * p1_amplitude')
    params.add('p2_sigma', expr='1 * p1_sigma')
    params.add('p2_fraction', expr='1 * p1_fraction')

    # Create a combined model by summing the two models
    combined_model = model1 + model2

    # Perform the fit
    fit_result = combined_model.fit(y, params, x=x)

    # Extract the fitted parameters for both peaks
    amplitude1 = fit_result.params['p1_amplitude'].value
    center1 = fit_result.params['p1_center'].value
    sigma1 = fit_result.params['p1_sigma'].value
    fraction1 = fit_result.params['p1_fraction'].value

    amplitude2 = fit_result.params['p2_amplitude'].value
    center2 = fit_result.params['p2_center'].value
    sigma2 = fit_result.params['p2_sigma'].value
    fraction2 = fit_result.params['p2_fraction'].value

    # Calculate FWHM for both peaks
    gamma1 = sigma1 / np.sqrt(2 * np.log(2))  # Convert sigma to gamma for Gaussian part
    fwhm1 = (1 - fraction1) * (2 * gamma1) + fraction1 * (2 * sigma1)

    gamma2 = sigma2 / np.sqrt(2 * np.log(2))
    fwhm2 = (1 - fraction2) * (2 * gamma2) + fraction2 * (2 * sigma2)

    return fit_result, amplitude1, fwhm1, center1, fraction1, amplitude2, fwhm2, center2, fraction2



def fit_this_peak(data, peak_position, fit_range, withplots = True, printinfo = False):

    cut_range = fit_range
    peak_angle = peak_position

    dat_theta = data.iloc[:,data.columns.get_level_values(1)=='2θ (°)']
    dat_counts = data.iloc[:,data.columns.get_level_values(1)=='Corrected Intensity']

    colors = plt.cm.jet(np.linspace(0, 1, len(dat_theta.columns)))

    plt.figure(figsize=(8, 6))

    df_fitted_peak = pd.DataFrame()

    for i in range(0, len(dat_theta.columns)):
        data_to_fit_x = dat_theta[dat_theta.columns[i]]
        data_to_fit_y = dat_counts[dat_counts.columns[i]]

        idx = np.where((data_to_fit_x >= peak_angle-cut_range) & (data_to_fit_x<=  peak_angle+cut_range))[0]
        x_range = data_to_fit_x[idx].values
        y_range = data_to_fit_y[idx].values

        fit_result, amplitude1, fwhm1, center1, fraction1, amplitude2, fwhm2, center2, fraction2 = fit_two_related_peaks(x_range, y_range)

        if printinfo == True:
            print(dat_theta.columns[i][0])
            print(f"Peak 1 - Amplitude: {amplitude1:.2f}, FWHM: {fwhm1:.2f}, Center: {center1:.2f}, Fraction: {fraction1:.2f}")
            print(f"Peak 2 - Amplitude: {amplitude2:.2f}, FWHM: {fwhm2:.2f}, Center: {center2:.2f}, Fraction: {fraction2:.2f}")

        if withplots==True:
            plt.plot(x_range, y_range, 'o', color = colors[i], label=str(dat_theta.columns[i][0]))
            plt.plot(x_range, fit_result.best_fit, '-', color = colors[i])
            plt.xlabel('2θ')
            plt.ylabel('Intensity')
            plt.title(' Fit with two related PseudoVoigts at '+ str(peak_angle) + '°')

        # store the information about the peak in a new dataframe 

        peakData = np.vstack((center1, amplitude1, fwhm1,  fraction1)).T
        peak_header = pd.MultiIndex.from_product([[dat_theta.columns[i][0]], ["Center","Amplitude", "FWHM", "Fraction"]], names = ["Coordinate", "Data type"])
        df_peak_info=pd.DataFrame( data= peakData, columns = peak_header)
        fitData = np.vstack((x_range, y_range, fit_result.best_fit)).T
        fit_header = pd.MultiIndex.from_product([[dat_theta.columns[i][0]], ["range 2θ","range Intensity", "Fit"]], names = ["Coordinate", "Data type"])
        df_fit_info = pd.DataFrame(data = fitData, columns = fit_header)
        df_fitted_peak = pd.concat([df_fitted_peak, df_fit_info, df_peak_info], axis=1)

    plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
    display(df_fitted_peak)
    return df_fitted_peak



def rgba_to_hex(rgba):
    """Convert an RGBA tuple to a hex color string."""
    r, g, b, a = [int(c * 255) for c in rgba]
    return f'#{r:02x}{g:02x}{b:02x}'

def interactive_XRD_shift(data, datatype_x, datatype_y, shift, x, y, ref_peaks_df, ref_label="Reference", title=None):
    'interactive shifted plot for assigning phases to XRD data'
    fig = make_subplots(
        rows=2, 
        cols=1, 
        shared_xaxes=True, 
        row_heights=[0.8, 0.2],  # Proportion of height for each plot
        vertical_spacing=0.02    # Adjust this to reduce space between plots
    )

    x_data = []
    y_data = []
    #colors = ['#636EFA', '#EF553B', '#00CC96', '#AB63FA', '#FFA15A']
    colormap = plt.get_cmap('turbo')  # You can choose any matplotlib colormap
    colors = [rgba_to_hex(colormap(i / len(x))) for i in range(len(x))]  # Convert colors to hex

    # Store all y-data to find the global maximum
    all_y_data = []

    # Loop through and plot the XRD spectra with a vertical shift in the top plot
    for i in range(len(x)):
        x_data = get_data(data, datatype_x, x[i], y[i], False, False)
        y_data = get_data(data, datatype_y, x[i], y[i], False, False)
        shifted_y_data = y_data - shift * i
        
        all_y_data.extend(shifted_y_data)  # Collect y-data with shift for max computation
        
        fig.add_trace(
            go.Scatter(
                x=x_data,
                y=shifted_y_data,
                mode='lines',
                line=dict(color=colors[i]),
                name=f'{x[i]}, {y[i]}'
            ),
            row=1, col=1
        )

    # Compute the global maximum y-value, considering shifts
    global_min_y = min(all_y_data)

    # Create traces for each reference material (hidden initially)
    ref_traces = []
    buttons = []

    for ref_material, ref_df in ref_peaks_df.items():
        # Reference spectrum plotted in the bottom plot
        ref_trace = go.Scatter(
            x=ref_df["2theta"],
            y=ref_df["I"],
            mode='lines',
            name=f'{ref_material} Reference',
            visible=False
        )
        
        # Create vertical peak lines for top plot (raw data plot)
        peak_lines = go.Scatter(
            x=[value for peak in ref_df["Peak 2theta"] for value in [peak, peak, None]],  # x: peak, peak, None to break the line
            y=[global_min_y, 1000 * 1.1, None] * len(ref_df["Peak 2theta"]),  # y: 0 -> global_max_y for each line, with None to break lines
            mode='lines',
            line=dict(color='grey', dash='dot'),
            showlegend=False,
            visible=False
        )

        # Append traces for each reference spectrum and its peaks
        ref_traces.append(ref_trace)
        ref_traces.append(peak_lines)
        
        # Create a button for each reference
        buttons.append(dict(
            label=ref_material,
            method='update',
            args=[{'visible': [True] * len(x) + [False] * len(ref_traces)},  # Show all raw spectra, hide refs by default
                  {'title': f'{title} - {ref_material} Reference'}]
        ))

    # Add reference traces to figure (initially hidden)
    for trace in ref_traces:
        # Ensure trace.name is not None before checking 'Reference' in name
        fig.add_trace(trace, row=2 if trace.name and 'Reference' in trace.name else 1, col=1)

    # Update buttons to control the visibility of one reference at a time
    for i, button in enumerate(buttons):
        # Make the selected reference spectrum visible in the bottom plot and its peaks visible in the top plot
        button['args'][0]['visible'][len(x):] = [False] * len(ref_traces)  # Hide all refs initially
        button['args'][0]['visible'][len(x) + 2 * i:len(x) + 2 * i + 2] = [True, True]  # Show selected ref and peaks

    # Add the dropdown menu to switch between reference spectra
    fig.update_layout(
        updatemenus=[{
            'buttons': buttons,
            'direction': 'down',
            'showactive': True,
            'x': 1.05,
            'xanchor': 'left',
            'y': 1.1,
            'yanchor': 'top'
        }],
        template='plotly_white',  # Choose a template (e.g., 'plotly_dark')
        title=title,
        height=600,  # Adjust the height of the figure (e.g., 700)
        width=900,   # Adjust the width of the figure (e.g., 900)
        legend=dict(x=1.05, y=1),
        xaxis2_title=datatype_x,
        yaxis_title=datatype_y
    )


    return fig

def extract_coordinates(data):
    coords= data.columns.get_level_values(0).unique().values
    x_values = []
    y_values = []
    
    for item in coords:
        x, y = item.split(',')
        x_values.append(float(x))
        y_values.append(float(y))
    
    return x_values, y_values


def EDX_stage_coords(folder, filename):
    'Calculate EDX coordinates for a given file. Requires .xlsx file with columns as in template.'

    # Define file paths
    filepath = os.path.join(folder, filename + ".xlsx")
    newpath = os.path.join(folder, filename + "_stage_coords.xlsx")
    
    # Read specific columns from the Excel file, and extract values
    file = pd.read_excel(filepath, sheet_name='Sheet2', usecols='H:P')
    file = file.drop(file.index[3:])  # Drop the 4th row (assumes you have info for 3 points)

    nrows = file['nrows'].values[0].astype(int)
    ncolumns = file['ncolumns'].values[0].astype(int)
    points_x = file['points x'].values[0:3]
    points_y = file['points y'].values[0:3]

    # Calculate spacing between points 1, 3, (nrows-1)
    space_x = points_x[2] - points_x[0]
    space_y = points_y[1] - points_y[0]

    # Generate coordinates and order them as layerprobe does
    coord_x = np.round(np.linspace(points_x[0], points_x[0] + space_x * (ncolumns - 1), ncolumns), 2)
    coord_y = np.round(np.linspace(points_y[0], points_y[0] + space_y * (nrows - 1), nrows), 2)

    X, Y = [], []
    for j in range(ncolumns):
        for i in range(nrows):
            Y.append(coord_y[i])
            X.append(coord_x[j])

    # Load the workbook and insert the calculated coordinates in the first sheet
    workbook = load_workbook(filepath)
    
    for i, value in enumerate(X, start=2):
        workbook['Sheet1'][f'B{i}'] = value
    for i, value in enumerate(Y, start=2):
        workbook['Sheet1'][f'C{i}'] = value

    workbook.save(newpath)
    workbook.close()
    
    print(filename, " - coordinates calculated and saved")



def EDX_sample_coords(folder, filename):
    'Calculate and translate EDX coordinates for a given file. Requires .xlsx file with columns as in template.'

    # Define file paths
    filepath = os.path.join(folder, filename+".xlsx")
    newpath = os.path.join(folder, filename+"_sample_coords.xlsx")

    # Read specific columns from the Excel file, and extract values
    file = pd.read_excel(filepath, sheet_name='Sheet2', usecols='H:P')
    file = file.drop(file.index[3:])  # Drop the 4th row (assumes you have info for 3 points)

    nrows= file['nrows'].values[0].astype(int)
    ncolumns = file['ncolumns'].values[0].astype(int)
    corners_x= file['corner x'].values[0:2]
    corners_y= file['corner y'].values[0:2]
    points_x = file['points x'].values[0:3]
    points_y = file['points y'].values[0:3]

    # Calculate spacing between points 1, 3, (nrows-1)
    space_x = points_x[2] - points_x[0]
    space_y = points_y[1] - points_y[0]

    # Calculate shift from corners and correct for this translation
    shift_x= (corners_x[1] +corners_x[0])/2
    shift_y= (corners_y[0] +corners_y[1])/2
    start_x = points_x[0] - shift_x
    start_y = points_y[0] - shift_y

    # Generate coordinates and order them as layerprobe does
    coord_x = np.round(np.linspace(start_x, start_x+ space_x*(ncolumns-1), ncolumns), 2)
    coord_y = np.round(np.linspace(start_y, start_y+ space_y*(nrows-1),    nrows), 2)
    X,Y = [],[]
    for j in range(0, ncolumns):
        for i in range(0, nrows):
            Y.append(coord_y[i])
            X.append(coord_x[j])

    # Load the workbook and insert the calculated coordinates in the first sheet
    workbook = load_workbook(filepath)
    sheet1= workbook['Sheet1']
    
    for i, value in enumerate(X, start= 2):
        sheet1[f'B{i}']= value
    for i, value in enumerate(Y, start= 2):
        sheet1[f'C{i}']= value

    workbook.save(newpath)
    workbook.close()

    # check for correct translation (allow 0.3 mm misalignment from sample rotation)
    if np.abs(X[-1]+X[0]) > 0.3 or np.abs(Y[-1]+Y[0]) > 0.3:
        print(filename, " - coordinates calculated and saved, but not symmetric")
        print("X shift: ", X[-1]+X[0])
        print("Y shift: ", Y[-1]+Y[0])
    else:
        print(filename, " - coordinates calculated, translated and saved")



def EDX_coordinates(folder, filename, edge=3, rotate=False, spacing= "auto"):
    'Calculate and translate EDX coordinates for a given file. Requires .xlsx file with columns as in template.'
    
    # Define file paths
    filepath = os.path.join(folder, filename+".xlsx")
    newpath = os.path.join(folder, filename+"_new_coords.xlsx")

    # Read specific columns from the Excel file, and extract values
    file = pd.read_excel(filepath, sheet_name='Sheet2', usecols='H:P')
    file = file.drop(file.index[3:])  # Drop the 4th row (assumes you have info for 3 points)

    nrows= file['nrows'].values[0].astype(int)
    ncolumns = file['ncolumns'].values[0].astype(int)
    corners_x= file['corner x'].values[0:2]
    corners_y= file['corner y'].values[0:2]
    mag = file['magnification'].values[0]
    if spacing == "auto":
        spacing = file['spacing'].values[0]

    if rotate == "90":
        ncolumns, nrows = nrows, ncolumns
        areax = 2.8 * 100 / mag
        areay = 4.1 * 100 / mag
    
    if rotate ==False:
        # Calculate the effective area size in the x-direction, considering magnification, 
        # assuming the x/y ratio is constant 4.1 : 2.8
        areax = 4.1 * 100 / mag
        areay= 2.8*100 / mag
    
    # Calculate the spacing , gridlength and starting x-coordinate for the grid in x-direction (assuming the grid is centered)
    space_x = areax * spacing / 100
    gridlength = (ncolumns - 1) * (space_x + areax) #+ areax
    startx = -gridlength / 2 #+ (areax / 2)

    # do the same for the y-direction
    
    space_y = areay*spacing/100
    gridheight= (nrows-1)*(space_y+ areay) #+ areay/2
    starty = -gridheight/2 #+ (areay/2)

    samplesize = [corners_x[1]-corners_x[0], corners_y[0]-corners_y[1]]
    print("Sample size is", samplesize)

    # Check if the grid dimensions exceed the maximum allowed size (31x31 mm)
    # if so, reduce the spacing by 10% and try again 
    if gridlength >= samplesize[0]-edge or gridheight >= samplesize[1]-edge:
        print("Spacing is too large for the map")

        new_spacing = np.round(spacing - spacing * 0.05, 0)
        print("New spacing is", new_spacing)

        return EDX_coordinates(folder, filename, spacing=new_spacing)

    # Generate coordinates for each column
    coord_x = np.round(np.linspace(startx, -startx, ncolumns), 2)
    coord_y = np.round(np.linspace(starty, -starty, nrows), 2)
    X=[]
    Y=[]
    for j in range(0, ncolumns):
        for i in range(0, nrows):
            Y.append(coord_y[i])
            X.append(coord_x[j])

    # Load the workbook and insert the calculated coordinates in the first sheet
    workbook = load_workbook(filepath)
    sheet1= workbook['Sheet1']
    
    for i, value in enumerate(X, start= 2):
        sheet1[f'B{i}']= value
    for i, value in enumerate(Y, start= 2):
        sheet1[f'C{i}']= value

    workbook.save(newpath)
    workbook.close()


def select_points(data, x_min=-18, x_max=18, y_min=-18, y_max=18):
    'get coordinates of the points within the defined range, you can call them with get_data, or plot_data, or interactive_XRD_shift'
    grid0 = MI_to_grid(data)
    grid = grid0.drop_duplicates().reset_index(drop=True)

    grid1 = grid[grid['x'] >x_min]
    grid2 = grid1[grid1['x'] <x_max]
    grid3 = grid2[grid2['y'] >y_min]
    grid4 = grid3[grid3['y'] <y_max]
    new_x = grid4['x'].values
    new_y = grid4['y'].values
    return new_x, new_y


def ternary_plot(df, el1, el2, el3, datatype, title, savepath = None): 
    "make a ternary plot of the data in df, with el1, el2, el3 as the corners, and colorscale based on datatype"

    A= f'Layer 1 {el1} Atomic %'
    B= f'Layer 1 {el2} Atomic %'
    C= f'Layer 1 {el3} Atomic %'

    A_percent = df.xs(A, level='Data type', axis=1).values.flatten()
    B_percent = df.xs(B, level='Data type', axis=1).values.flatten()
    C_percent = df.xs(C, level='Data type', axis=1).values.flatten()
    intensity = df.xs(datatype, level='Data type', axis=1).values.flatten()
    coordinates = MI_to_grid(df).values # df.columns.get_level_values(0)[::7].values.flatten() also works, but gives a lot of decimals

    fig = go.Figure()
    fig.add_trace(go.Scatterternary({
        'mode': 'markers',
        'a': A_percent,   # el1 percentages
        'b': B_percent,   # el2 percentages
        'c': C_percent,   # el3 percentages
        'marker': {
            'symbol': 100 ,
            'size': 8,
            'color': intensity,  # Use intensity for marker color
            'colorscale': 'Turbo',  # Choose a colorscale
            'colorbar': {'title': datatype},  # Add a colorbar
            'line': {'width': 2}
        },
        'text': coordinates,
        'hovertemplate': f'{el1}: %{{a:.1f}}%<br>{el2}: %{{b:.1f}}%<br>{el3}: %{{c:.1f}}%<br>{datatype}: %{{marker.color:.1f}}<br>Coordinates:%{{text:str}}',  # Custom hover text format
        #'name': f'{datatype}: %{{text:str}}'
        'showlegend': False
    }))

    # Update layout
    fig.update_layout({
        'ternary': {
            'sum': 100,
            'aaxis': {'title': f'{el1} %', 'min': 0, 'linewidth': 2, 'ticks': 'outside'},
            'baxis': {'title': f'{el2} %', 'min': 0, 'linewidth': 2, 'ticks': 'outside'},
            'caxis': {'title': f'{el3} %', 'min': 0, 'linewidth': 2, 'ticks': 'outside'}
        },
        'title': title},
        width=800,
        height=600, 
    )       
    if savepath:
        if savepath.endswith(".png"):
            fig.write_image(savepath, scale=2)
        if savepath.endswith(".html"):
            fig.write_html(savepath)

    # Show the plot
    fig.show()

# %%
